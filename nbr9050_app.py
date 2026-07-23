"""
♿ Auditor de Acessibilidade BIM — NBR 9050:2020
Streamlit App — Verificação Automatizada de Conformidade
Master Internacional em IA para Arquitetura e Construção — Zigurat Institute of Technology
"""

import streamlit as st
import json
import os
import re
import time
import tempfile
import io
from pathlib import Path
from datetime import datetime

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Auditor NBR 9050 — BIM",
    page_icon="♿",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>

:root {
    --zk-slate:  rgb(77,83,99);
    --zk-green:  rgb(68,205,148);
    --zk-blue:   rgb(28,96,241);
    --bg:        #ffffff;
    --surface:   #f4f6f9;
    --border:    #d1d5de;
    --border-dk: rgb(77,83,99);
    --text:      #1a1d26;
    --muted:     #6b7280;
    --success:   #1ab87a;
    --danger:    #e03c3c;
    --warn:      #e8920a;
    --font: 'Trebuchet MS', Trebuchet, Arial, sans-serif;
    --mono: 'Courier New', Courier, monospace;
}

/* ── Global ── */
html, body, .stApp, [class*="css"] {
    font-family: var(--font) !important;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}
.stApp { background-color: var(--bg) !important; }
.stApp p, .stApp span, .stApp label, .stApp div,
.stApp h1, .stApp h2, .stApp h3, .stApp li { color: var(--text) !important; }

/* ════════════════════════════════════
   SIDEBAR — fundo BRANCO
   ════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 5px solid var(--zk-green) !important;
}
/* Textos escuros na sidebar branca */
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] div,
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] .stMarkdown { color: var(--text) !important; }

/* Inputs na sidebar: fundo cinza claro, texto escuro */
[data-testid="stSidebar"] input {
    background: var(--surface) !important;
    color: var(--text) !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
    font-family: var(--font) !important;
}
[data-testid="stSidebar"] input::placeholder { color: var(--muted) !important; }

/* Selectbox na sidebar */
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
}
[data-testid="stSidebar"] [data-baseweb="select"] span,
[data-testid="stSidebar"] [data-baseweb="select"] div {
    color: var(--text) !important;
    font-family: var(--font) !important;
}

/* Dropdown list */
[data-baseweb="popover"] [role="option"] {
    background: #ffffff !important;
    color: var(--text) !important;
    font-family: var(--font) !important;
}
[data-baseweb="popover"] [role="option"]:hover,
[data-baseweb="popover"] [aria-selected="true"] {
    background: rgba(68,205,148,0.15) !important;
    color: var(--text) !important;
}

/* Linha separadora verde na sidebar */
[data-testid="stSidebar"] hr { border-color: var(--zk-green) !important; border-width: 2px !important; }

/* Labels da sidebar */
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stSlider label { color: var(--zk-slate) !important; font-weight: 600 !important; }

/* ── Slider na sidebar ── */
[data-testid="stSidebar"] [data-baseweb="slider"] { background: var(--border) !important; }
[data-testid="stSidebar"] [role="slider"] {
    background: var(--zk-green) !important;
    width: 20px !important; height: 20px !important;
    border: 3px solid #ffffff !important;
    box-shadow: 0 0 0 2px var(--zk-green) !important;
}
[data-testid="stSidebar"] [data-testid="stSlider"] [data-baseweb="slider"] div[role="progressbar"] {
    background: var(--zk-green) !important;
}
/* Ocultar o valor flutuante nativo (tooltip do thumb) */
[data-testid="stSidebar"] [data-testid="stSlider"] [data-baseweb="tooltip"],
[data-testid="stSidebar"] [data-testid="stSlider"] div[data-testid="stTickBarMin"],
[data-testid="stSidebar"] [data-testid="stSlider"] div[data-testid="stTickBarMax"] {
    display: none !important;
}

/* ════════════════════════════════════
   HERO — fundo BRANCO, borda escura
   ════════════════════════════════════ */
.hero-block {
    background: #ffffff;
    border: 2px solid var(--zk-slate);
    border-radius: 10px;
    padding: 1.5rem 2rem;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 1rem;
}
.hero-title {
    font-family: var(--font);
    font-size: 1.75rem;
    font-weight: 700;
    color: var(--zk-slate) !important;
    margin: 0 0 0.2rem 0;
}
.hero-sub {
    font-family: var(--font);
    font-size: 0.72rem;
    color: var(--muted) !important;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}

/* ── Section titles ── */
.section-title {
    font-family: var(--font);
    font-size: 0.78rem;
    font-weight: 700;
    color: var(--zk-blue) !important;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    margin-bottom: 0.85rem;
    display: flex;
    align-items: center;
    gap: 0.4rem;
}

/* ── Status badges ── */
.badge { display: inline-block; padding: 2px 10px; border-radius: 20px; font-family: var(--font); font-size: 0.72rem; font-weight: 600; }
.badge-conforme { background: rgba(26,184,122,0.12); color: #0a7a4e !important; border: 1px solid rgba(26,184,122,0.35); }
.badge-parcial  { background: rgba(124,58,196,0.12);  color: #5b21a6 !important; border: 1px solid rgba(124,58,196,0.35); }
.badge-nao      { background: rgba(224,60,60,0.10);  color: #9b1c1c !important; border: 1px solid rgba(224,60,60,0.35); }
.badge-indet    { background: rgba(232,146,10,0.12); color: #7a4500 !important; border: 1px solid rgba(232,146,10,0.35); }
.badge-na       { background: rgba(77,83,99,0.08);   color: #4d5363 !important; border: 1px solid rgba(77,83,99,0.2); }

/* ── Result table ── */
.result-table { width: 100%; border-collapse: collapse; font-size: 0.81rem; }
.result-table th {
    font-family: var(--font); font-size: 0.67rem; color: #ffffff !important;
    text-transform: uppercase; letter-spacing: 0.08em;
    border-bottom: 2px solid var(--zk-blue); padding: 9px 12px;
    text-align: left; background: var(--zk-slate);
}
.result-table td { padding: 9px 12px; border-bottom: 1px solid var(--border); vertical-align: top; color: var(--text) !important; }
.result-table tr:hover td { background: rgba(68,205,148,0.05); }

/* ── GlobalId chip ── */
.globalid { font-family: var(--mono); font-size: 0.67rem; background: #eef1f8; border: 1px solid #c5cad8; border-radius: 4px; padding: 2px 6px; color: var(--zk-blue) !important; display: inline-block; }

/* ── Metric cards ── */
.metric-row { display: flex; gap: 0.85rem; margin-bottom: 1.5rem; flex-wrap: wrap; }
.metric-card { flex: 1; min-width: 110px; background: var(--surface); border: 1px solid var(--border); border-top: 3px solid var(--zk-green); border-radius: 8px; padding: 1rem; text-align: center; }
.metric-num { font-family: var(--font); font-size: 2rem; font-weight: 700; line-height: 1; margin-bottom: 0.2rem; }
.metric-label { font-family: var(--font); font-size: 0.62rem; color: var(--muted) !important; text-transform: uppercase; letter-spacing: 0.08em; }
.c-green { color: var(--success) !important; }
.c-purple{ color: #7c3ac4 !important; }
.c-red   { color: var(--danger)  !important; }
.c-amber { color: var(--warn)    !important; }
.c-blue  { color: var(--zk-blue) !important; }
.c-muted { color: var(--muted)   !important; }

/* ── Terminal ── */
.terminal {
    background: var(--zk-slate); border: 1px solid #3a3f4f; border-radius: 8px;
    padding: 1rem 1.25rem; font-family: var(--mono); font-size: 0.74rem;
    color: var(--zk-green) !important; max-height: 280px; overflow-y: auto; line-height: 1.7;
}

/* ── Inputs main content ── */
.stTextInput input, .stTextArea textarea {
    background: var(--surface) !important; border: 1px solid var(--border) !important;
    color: var(--text) !important; border-radius: 6px !important;
    font-family: var(--font) !important; font-size: 0.82rem !important;
}
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color: var(--zk-green) !important;
    box-shadow: 0 0 0 2px rgba(68,205,148,0.2) !important;
}

/* Selectbox main content */
[data-testid="stMain"] [data-baseweb="select"] > div { background: var(--surface) !important; border: 1px solid var(--border) !important; color: var(--text) !important; }
[data-testid="stMain"] [data-baseweb="select"] span { color: var(--text) !important; }

/* ── File uploader — borda verde chamativa ── */
[data-testid="stFileUploader"] { background: #f0fdf8 !important; border: 2px dashed var(--zk-green) !important; border-radius: 8px !important; }
[data-testid="stFileUploader"]:hover { background: #e6faf3 !important; border-color: var(--zk-blue) !important; }
[data-testid="stFileUploader"] span, [data-testid="stFileUploader"] p { color: var(--text) !important; }

/* ── Botão principal — verde chamativo ── */
.stButton button {
    background: var(--zk-green) !important; color: rgb(20,60,40) !important;
    font-family: var(--font) !important; font-weight: 700 !important; font-size: 0.9rem !important;
    border: none !important; border-radius: 6px !important; padding: 0.65rem 1.75rem !important;
    transition: all 0.2s !important; letter-spacing: 0.02em !important;
}
.stButton button:hover { background: var(--zk-blue) !important; color: #ffffff !important; transform: translateY(-1px); box-shadow: 0 4px 18px rgba(28,96,241,0.3) !important; }
.stButton button:disabled { background: var(--border) !important; color: var(--muted) !important; transform: none !important; }

/* ── Download buttons ── */
[data-testid="stDownloadButton"] button { background: var(--surface) !important; color: var(--zk-blue) !important; border: 1.5px solid var(--zk-blue) !important; font-family: var(--font) !important; font-weight: 600 !important; }
[data-testid="stDownloadButton"] button:hover { background: var(--zk-blue) !important; color: #ffffff !important; }

/* ── Expander ── */
.streamlit-expanderHeader { background: var(--surface) !important; border: 1px solid var(--border) !important; border-radius: 6px !important; font-family: var(--font) !important; font-size: 0.82rem !important; color: var(--text) !important; }

/* ── Tabs ── */
[data-testid="stTabs"] [role="tab"] { font-family: var(--font) !important; font-size: 0.82rem !important; color: var(--muted) !important; }
[data-testid="stTabs"] [role="tab"][aria-selected="true"] { color: var(--zk-blue) !important; border-bottom-color: var(--zk-blue) !important; }

/* ── Info / warning boxes ── */
.info-box { background: rgba(28,96,241,0.06); border-left: 3px solid var(--zk-blue); border-radius: 0 6px 6px 0; padding: 0.75rem 1rem; font-size: 0.82rem; color: #1a3fa8 !important; margin: 0.75rem 0; }
.warn-box { background: rgba(232,146,10,0.08); border-left: 3px solid var(--warn); border-radius: 0 6px 6px 0; padding: 0.75rem 1rem; font-size: 0.82rem; color: #7a4500 !important; margin: 0.75rem 0; }

/* ── Divider ── */
hr { border-color: var(--border) !important; }

/* ── Pulse ── */
@keyframes pulse { 0%,100% { box-shadow: 0 0 0 0 rgba(68,205,148,0.4); } 50% { box-shadow: 0 0 0 8px rgba(68,205,148,0); } }

/* ── Botão CTA — verde vibrante com sombra quando habilitado ── */
.stButton button {
    background: var(--zk-green) !important; color: rgb(20,60,40) !important;
    font-family: var(--font) !important; font-weight: 700 !important; font-size: 0.9rem !important;
    border: none !important; border-radius: 6px !important; padding: 0.65rem 1.75rem !important;
    transition: all 0.2s !important; letter-spacing: 0.02em !important;
    box-shadow: 0 2px 8px rgba(68,205,148,0.35) !important;
}
.stButton button:hover {
    background: var(--zk-blue) !important; color: #ffffff !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 20px rgba(28,96,241,0.35) !important;
}
.stButton button:disabled {
    background: #e5e7eb !important; color: #9ca3af !important;
    transform: none !important; box-shadow: none !important;
}

/* ════════════════════════════════════
   RODAPÉ — fundo BRANCO, borda escura
   ════════════════════════════════════ */
.footer-app {
    background: #ffffff;
    border: 2px solid var(--zk-slate);
    border-radius: 8px;
    padding: 1rem 1.5rem;
    margin-top: 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 0.75rem;
}
.footer-app-tfm { color: var(--zk-green) !important; font-weight: 700; font-size: 0.72rem; margin-bottom: 0.15rem; }
.footer-app-autores { color: var(--zk-slate) !important; font-size: 0.65rem; }

/* ── Ocultar chrome do Streamlit ── */
#MainMenu, footer, header { visibility: hidden; }

/* ── Manter sidebar sempre visível ── */
[data-testid="collapsedControl"] { display: none !important; }
section[data-testid="stSidebar"] { min-width: 240px !important; transform: none !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _estatisticas_portas(portas_todas: list[dict]) -> dict:
    """
    Calcula min/max e contagem de não conformidades sobre TODAS as portas
    do modelo — não só a amostra de 60 que é enviada ao prompt.

    POR QUÊ ISSO EXISTE: a amostra (`portas[:60]`) existe pra não estourar
    o orçamento de tokens do prompt quando o modelo tem centenas de portas.
    Mas se o LLM só vê 60 de, digamos, 265 portas, ele só consegue avaliar
    conformidade das 60 — as outras 205 nunca são checadas, e uma porta fora
    do padrão nelas passaria batido. Esta função varre a lista COMPLETA em
    Python (rápido, determinístico) e devolve um resumo estatístico que é
    anexado ao prompt ao lado da amostra — assim o veredito do LLM cobre
    o modelo inteiro, não só a fatia que caiu na amostra ilustrativa.
    """
    larguras = [(d["GlobalId"], d["OverallWidth_m"]) for d in portas_todas if d.get("OverallWidth_m") is not None]
    alturas  = [(d["GlobalId"], d["OverallHeight_m"]) for d in portas_todas if d.get("OverallHeight_m") is not None]

    if not larguras and not alturas:
        return {"total": len(portas_todas), "com_dimensoes": 0}

    gid_min_larg, min_larg = min(larguras, key=lambda x: x[1]) if larguras else (None, None)
    gid_min_alt, min_alt   = min(alturas, key=lambda x: x[1]) if alturas else (None, None)

    # NBR 9050 item 6.11.2: largura ≥ 0,80m e altura ≥ 2,10m
    nao_conf_largura = [gid for gid, l in larguras if l < 0.80]
    nao_conf_altura  = [gid for gid, a in alturas if a < 2.10]

    return {
        "total": len(portas_todas),
        "com_dimensoes": len(set(gid for gid, _ in larguras) | set(gid for gid, _ in alturas)),
        "largura_min_m": round(min_larg, 3) if min_larg is not None else None,
        "largura_min_globalid": gid_min_larg,
        "altura_min_m": round(min_alt, 3) if min_alt is not None else None,
        "altura_min_globalid": gid_min_alt,
        "n_nao_conformes_largura": len(nao_conf_largura),
        "n_nao_conformes_altura": len(nao_conf_altura),
        "globalids_nao_conformes_largura": nao_conf_largura[:15],
        "globalids_nao_conformes_altura": nao_conf_altura[:15],
    }


def extract_ifc_elements(ifc_path: str) -> dict:
    """
    Extrai e pré-processa elementos do IFC para auditoria NBR 9050.
    
    DESCOBERTAS DO MODELO AGO-ARQ (IFC2X3/Revit):
    - IfcStairFlight: em IFC2X3, NumberOfRisers/RiserHeight/TreadLength são atributos
      diretos da entidade e vêm em PÉS (→ x0.3048). Em IFC4 esses dados saem do
      Pset_StairFlightCommon/Pset_StairCommon ("NumberOfRiser" no singular) já em
      METROS — extract_ifc_elements detecta qual caso se aplica por elemento.
    - IfcDoor: OverallHeight (pos 9), OverallWidth (pos 10) estão em METROS → correto
    - IfcFlowTerminal: contém bacias, lavatórios, barras de apoio, torneiras — filtrar por nome
    - IfcRailing: contém guarda-corpos (não corrimão) — modelo não tem corrimão separado
    - IfcSpace: ausente neste modelo — fallback via IfcWall necessário
    - IfcRamp/IfcRampFlight: ausentes neste modelo
    """
    try:
        import ifcopenshell
    except ImportError:
        return {"error": "ifcopenshell não instalado. Execute: pip install ifcopenshell"}

    ifc    = ifcopenshell.open(ifc_path)
    schema = ifc.schema

    FT_TO_M = 0.3048  # Revit exporta dimensões de escadas em pés para IFC2X3

    def todos_psets(el):
        psets = {}
        try:
            for rel in getattr(el, "IsDefinedBy", []):
                if rel.is_a("IfcRelDefinesByProperties"):
                    pdef = rel.RelatingPropertyDefinition
                    if pdef.is_a("IfcPropertySet"):
                        props = {}
                        for p in getattr(pdef, "HasProperties", []):
                            try:
                                val = None
                                if hasattr(p, "NominalValue") and p.NominalValue:
                                    val = p.NominalValue.wrappedValue
                                props[p.Name] = val
                            except Exception:
                                pass
                        if props:
                            psets[pdef.Name] = props
        except Exception:
            pass
        return psets

    def info_basica(el, tipo_ifc=None):
        return {
            "GlobalId":    el.GlobalId,
            "Name":        getattr(el, "Name", None),
            "ObjectType":  getattr(el, "ObjectType", None),
            "Tag":         getattr(el, "Tag", None),
            "Description": getattr(el, "Description", None),
            "tipo_ifc":    tipo_ifc or el.is_a(),
        }

    def buscar_prop(psets, *termos):
        for ps in psets.values():
            for k, v in ps.items():
                if any(t in k.lower() for t in termos) and v is not None:
                    return v
        return None

    def buscar_todas_props(psets, *termos):
        """
        Como buscar_prop, mas coleta TODAS as ocorrências em vez de parar na
        primeira. Necessário para elementos compostos — ex: um único IfcRailing
        do Revit pode carregar 3 Psets diferentes ("Corrimão 1", "Corrimão 2",
        "Corrimão superior"), cada um com sua própria propriedade "Altura".
        Pegar só a primeira jogava fora 2 das 3 alturas reais do elemento.
        """
        achados = []
        for pset_nome, ps in psets.items():
            for k, v in ps.items():
                if any(t in k.lower() for t in termos) and v is not None:
                    achados.append({"pset": pset_nome, "propriedade": k, "valor": v})
        return achados

    def limpar_nulos(obj):
        if isinstance(obj, dict):
            return {k: limpar_nulos(v) for k, v in obj.items() if v is not None}
        if isinstance(obj, list):
            return [limpar_nulos(i) for i in obj if i is not None]
        return obj

    def contar(tipo):
        try: return len(ifc.by_type(tipo))
        except: return 0

    resultado = {"schema": schema, "arquivo": Path(ifc_path).name, "elementos": {}}

    # ── Inventário ────────────────────────────────────────────────────────────
    resultado["inventario_modelo"] = {
        t: contar(t) for t in [
            "IfcDoor","IfcWindow","IfcStair","IfcStairFlight","IfcRamp","IfcRampFlight",
            "IfcRailing","IfcSpace","IfcSlab","IfcWall","IfcWallStandardCase",
            "IfcFlowTerminal","IfcFurnishingElement","IfcBuildingElementProxy",
        ]
    }

    # ── Filtro PNE/PCD — usado em portas (4.6.6) e sanitários (7.x) ──────────
    # A NBR 9050 item 4.6.6 (maçaneta tipo alavanca) é uma exigência geral de
    # rota acessível, mas por decisão do projeto/thesis, aqui só é cobrada nas
    # portas de sanitários PNE/PCD — as demais portas do modelo ficam fora do
    # escopo dessa checagem específica. Verifica: (1) nome/tipo do próprio
    # elemento; (2) se ausente, nome do IfcSpace que o contém.
    TERMOS_ACESSIVEL = ["pne", "pcd"]
    try:
        from ifcopenshell.util.element import get_container as _get_container
    except Exception:
        _get_container = None

    def eh_acessivel_pne(el, texto_proprio):
        if any(t in texto_proprio for t in TERMOS_ACESSIVEL):
            return True, "nome_elemento"
        if _get_container:
            try:
                cont = _get_container(el)
                if cont is not None and cont.is_a("IfcSpace"):
                    nome_cont = ((getattr(cont, "Name", "") or "") + " " + (getattr(cont, "LongName", "") or "")).lower()
                    if any(t in nome_cont for t in TERMOS_ACESSIVEL):
                        return True, "nome_espaco_continente"
            except Exception:
                pass
        return False, None

    # ── 1. PORTAS (6.11.2, 4.6.6) ────────────────────────────────────────────
    # IfcDoor IFC2X3: campo 9=OverallHeight, campo 10=OverallWidth (em metros)
    portas = []
    for el in ifc.by_type("IfcDoor"):
        d = info_basica(el, "IfcDoor")
        oh = getattr(el, "OverallHeight", None)
        ow = getattr(el, "OverallWidth", None)
        # Em IFC2X3 Revit: OverallHeight é o 1º parâmetro dimensional, OverallWidth o 2º
        d["OverallHeight_m"] = round(float(oh), 3) if oh else None
        d["OverallWidth_m"]  = round(float(ow), 3) if ow else None
        d["Psets"] = todos_psets(el)

        nome_porta = (getattr(el, "Name", "") or "").lower() + " " + (getattr(el, "ObjectType", "") or "").lower()
        pne_ok, pne_fonte = eh_acessivel_pne(el, nome_porta)
        d["pne_pcd_confirmado"] = pne_ok  # relevante só pro item 4.6.6 (maçaneta) — 6.11.2 (vão livre) vale pra todas
        if pne_fonte:
            d["pne_pcd_fonte"] = pne_fonte

        portas.append(d)
    resultado["elementos"]["IfcDoor"] = portas[:60]
    resultado["estatisticas_portas"] = _estatisticas_portas(portas)
    n_portas_pne = sum(1 for p in portas if p.get("pne_pcd_confirmado"))
    resultado["nota_portas_pne"] = (
        f"{n_portas_pne} de {len(portas)} portas foram identificadas em ambiente/nome PNE/PCD. "
        f"O item 4.6.6 (maçaneta tipo alavanca) deve ser avaliado SOMENTE nessas portas — "
        f"as demais ficam fora do escopo desse item específico (mas continuam valendo para 6.11.2, vão livre)."
    )

    # ── 2. RAMPAS (6.6) ─────────────────────────────────────────────────────
    def _geom_bbox_rise_run(el):
        """
        Fallback geométrico: quando não há OverallRise/OverallRun nem no atributo
        direto nem no Pset, estima a partir da geometria bruta.

        run = maior dimensão em planta (X ou Y) — essa parte é confiável.

        rise = precisa de cuidado: a rampa é modelada como uma LAJE inclinada
        com espessura própria (não uma superfície fina). Pegar direto
        (Z_max global − Z_min global) do sólido inteiro conta a espessura da
        laje NAS DUAS PONTAS junto com o desnível real, inflando o resultado
        (confirmado num caso real: bbox bruto deu 0,51m onde o desnível real,
        medido ponta a ponta pela cota média de cada extremidade, era 0,36m —
        diferença suficiente pra trocar "Não Conforme" por "Conforme").
        Para evitar isso: agrupa os vértices pelas pontas ao longo do eixo de
        percurso (percentis 15/85) e compara a cota Z MÉDIA de cada ponta —
        isso cancela a espessura da laje, que aparece igualmente nas duas pontas.
        """
        try:
            import ifcopenshell.geom
            import numpy as np
            gset = ifcopenshell.geom.settings()
            gset.set(gset.USE_WORLD_COORDS, True)
            shape = ifcopenshell.geom.create_shape(gset, el)
            verts = np.array(shape.geometry.verts).reshape(-1, 3)

            x_range = verts[:, 0].max() - verts[:, 0].min()
            y_range = verts[:, 1].max() - verts[:, 1].min()
            axis_idx = 0 if x_range >= y_range else 1
            run = round(float(max(x_range, y_range)), 3)

            axis_vals = verts[:, axis_idx]
            lo, hi = np.percentile(axis_vals, [15, 85])
            grupo_inicio = verts[axis_vals <= lo]
            grupo_fim    = verts[axis_vals >= hi]
            if len(grupo_inicio) and len(grupo_fim):
                rise = round(float(abs(grupo_fim[:, 2].mean() - grupo_inicio[:, 2].mean())), 3)
            else:
                rise = round(float(verts[:, 2].max() - verts[:, 2].min()), 3)  # fallback bruto

            return rise, run
        except Exception:
            return None, None

    rampas = []
    for tipo in ["IfcRamp", "IfcRampFlight"]:
        for el in ifc.by_type(tipo):
            d = info_basica(el, tipo)
            ps = todos_psets(el)
            d["Psets"] = ps

            rise_attr = getattr(el, "OverallRise", None)
            run_attr  = getattr(el, "OverallRun", None)
            rise_pset = buscar_prop(ps, "overallrise", "altura da rampa", "desnivel")
            run_pset  = buscar_prop(ps, "overallrun", "comprimento da rampa")
            slope_pset = buscar_prop(ps, "slope", "inclinacao", "inclinação")

            rise_m, run_m, fonte = None, None, "nao_encontrado"
            if rise_attr and run_attr:
                # IFC2X3: atributo direto, em pés
                rise_m = round(float(rise_attr) * FT_TO_M, 3)
                run_m  = round(float(run_attr) * FT_TO_M, 3)
                fonte = "atributo_direto_ifc2x3_pes"
            elif rise_pset and run_pset:
                # IFC4: Pset_RampFlightCommon/RampCommon, já em metros
                rise_m = round(float(rise_pset), 3)
                run_m  = round(float(run_pset), 3)
                fonte = "pset_ifc4_metros"
            else:
                # Nenhum dos dois → estima pela geometria (bounding box)
                rise_geo, run_geo = _geom_bbox_rise_run(el)
                if rise_geo and run_geo:
                    rise_m, run_m = rise_geo, run_geo
                    fonte = "geometria_bounding_box_ESTIMATIVA"

            d["OverallRise_m"] = rise_m
            d["OverallRun_m"]  = run_m
            d["fonte_dados_rampa"] = fonte
            if slope_pset is not None:
                d["Slope_pset_bruto"] = slope_pset  # valor cru do Pset — unidade não confirmada, conferir
            if rise_m and run_m and run_m > 0:
                d["inclinacao_pct"] = round(rise_m / run_m * 100, 2)
            rampas.append(d)

    # Fallback: IfcSlab modelado como rampa (nome contém "rampa"/"ramp"/"slope")
    for el in ifc.by_type("IfcSlab"):
        nome = (getattr(el, "Name", "") or "").lower()
        otype = (getattr(el, "ObjectType", "") or "").lower()
        if any(t in nome + otype for t in ["rampa", "ramp", "slope"]):
            d = info_basica(el, "IfcSlab(rampa-fallback)")
            d["Psets"] = todos_psets(el)
            rise_geo, run_geo = _geom_bbox_rise_run(el)
            if rise_geo and run_geo:
                d["OverallRise_m"] = rise_geo
                d["OverallRun_m"] = run_geo
                d["fonte_dados_rampa"] = "geometria_bounding_box_slab_ESTIMATIVA"
                if run_geo > 0:
                    d["inclinacao_pct"] = round(rise_geo / run_geo * 100, 2)
            rampas.append(d)
    resultado["elementos"]["Rampas"] = rampas
    resultado["nota_rampas"] = f"Modelo tem {contar('IfcRamp')} IfcRamp e {contar('IfcRampFlight')} IfcRampFlight. Sem rampas modeladas neste projeto."

    # ── 3. ESCADAS + DESNÍVEL CALCULADO (5.4.3) ───────────────────────────────
    # ATENÇÃO: RiserHeight e TreadLength do Revit/IFC2X3 estão em PÉS → x0.3048
    escadas = []
    for el in ifc.by_type("IfcStairFlight"):
        d = info_basica(el, "IfcStairFlight")
        ps = todos_psets(el)
        d["Psets"] = ps

        # ── Estratégia por CAMPO (não por elemento inteiro): cada campo tenta o
        # atributo direto (IFC2X3, em pés) e, se vier vazio, cai pro Pset (IFC4,
        # "NumberOfRiser" no singular, já em metros) — independentemente dos
        # outros campos. Isso cobre o caso real de um Revit preencher RiserHeight
        # automaticamente mas deixar NumberOfRisers em branco (ou vice-versa),
        # que a versão anterior (cascata por elemento) não pegava.
        nr_attr = getattr(el, "NumberOfRisers", None)
        nt_attr = getattr(el, "NumberOfTreads", None)
        rh_attr = getattr(el, "RiserHeight", None)
        tl_attr = getattr(el, "TreadLength", None)

        fontes = []

        if nr_attr is not None:
            nr = nr_attr
            fontes.append("NumberOfRisers:atributo")
        else:
            nr = buscar_prop(ps, "numberofriser", "numberofrisers", "número de espelhos", "numero de espelhos")
            if nr is not None:
                fontes.append("NumberOfRisers:pset")

        if nt_attr is not None:
            nt = nt_attr
            fontes.append("NumberOfTreads:atributo")
        else:
            nt = buscar_prop(ps, "numberoftread", "número de pisos", "numero de pisos")
            if nt is not None:
                fontes.append("NumberOfTreads:pset")

        if rh_attr is not None:
            rh_m = round(float(rh_attr) * FT_TO_M, 4)
            fontes.append("RiserHeight:atributo_pes")
        else:
            rh_pset = buscar_prop(ps, "riserheight", "altura do espelho")
            rh_m = round(float(rh_pset), 4) if rh_pset is not None else None
            if rh_m is not None:
                fontes.append("RiserHeight:pset_metros")

        if tl_attr is not None:
            tl_m = round(float(tl_attr) * FT_TO_M, 4)
            fontes.append("TreadLength:atributo_pes")
        else:
            tl_pset = buscar_prop(ps, "treadlength", "largura do piso")
            tl_m = round(float(tl_pset), 4) if tl_pset is not None else None
            if tl_m is not None:
                fontes.append("TreadLength:pset_metros")

        d["fonte_dados_escada"] = ", ".join(fontes) if fontes else "nao_encontrado"

        d["NumberOfRisers"] = int(float(nr)) if nr is not None else None
        d["NumberOfTreads"] = int(float(nt)) if nt is not None else None
        d["RiserHeight_m"]  = rh_m
        d["TreadLength_m"]  = tl_m

        # Desnível calculado = NumberOfRisers × RiserHeight (já em metros)
        if d["NumberOfRisers"] and d["RiserHeight_m"]:
            d["desnivel_m"] = round(d["NumberOfRisers"] * d["RiserHeight_m"], 3)

        escadas.append(d)

    # Também inclui IfcStair (container) para contexto
    for el in ifc.by_type("IfcStair"):
        d = info_basica(el, "IfcStair")
        d["Psets"] = todos_psets(el)
        escadas.append(d)

    resultado["elementos"]["Escadas"] = escadas
    resultado["nota_escadas"] = (
        f"RiserHeight_m e TreadLength_m já normalizados para metros — via atributo direto "
        f"convertido de pés (schema IFC2X3) ou via Pset_StairCommon/Pset_StairFlightCommon já "
        f"em metros (schema IFC4). Veja 'fonte_dados_escada' em cada item para a origem. "
        f"Use desnivel_m = NumberOfRisers × RiserHeight_m para calcular o desnível total."
    )

    # ── 4. CORRIMÕES / GUARDA-CORPOS (5.4.3) ─────────────────────────────────
    TOLERANCIA_ALTURA = 0.03  # 3cm de tolerância pra bater com 0,70m/0,92m
    corrimaos = []
    algum_com_corrimao_duplo = False
    for el in ifc.by_type("IfcRailing"):
        d = info_basica(el, "IfcRailing")
        ps = todos_psets(el)
        d["Psets"] = ps

        # Um único IfcRailing pode conter VÁRIOS sub-Psets com "Altura" própria
        # (ex: corrimão inferior, corrimão superior, guarda-corpo) — coleta todas,
        # não só a maior/primeira encontrada.
        alturas_encontradas = buscar_todas_props(ps, "altura", "height")
        d["alturas_detalhadas"] = alturas_encontradas
        valores = sorted({round(float(a["valor"]), 3) for a in alturas_encontradas})
        d["alturas_m"] = valores

        tem_070 = any(abs(v - 0.70) <= TOLERANCIA_ALTURA for v in valores)
        tem_092 = any(abs(v - 0.92) <= TOLERANCIA_ALTURA for v in valores)
        d["corrimao_duplo_070_092"] = tem_070 and tem_092
        if d["corrimao_duplo_070_092"]:
            algum_com_corrimao_duplo = True

        nome = (getattr(el, "Name", "") or "").lower()
        d["tipo_elemento"] = "guarda-corpo" if "guarda" in nome else ("corrimao" if "corrim" in nome else "railing")
        corrimaos.append(d)

    resultado["elementos"]["Corrimaos"] = corrimaos
    if not corrimaos:
        resultado["nota_corrimaos"] = "Modelo não tem nenhum IfcRailing."
    elif algum_com_corrimao_duplo:
        resultado["nota_corrimaos"] = (
            f"Modelo tem {len(corrimaos)} IfcRailing. Pelo menos um contém, em Psets separados dentro do "
            f"MESMO elemento, alturas compatíveis com corrimão duplo (0,70m e 0,92m) — verifique o campo "
            f"'alturas_m' de cada item para ver todas as alturas detectadas por elemento."
        )
    else:
        alturas_todas = sorted({v for c in corrimaos for v in c.get("alturas_m", [])})
        resultado["nota_corrimaos"] = (
            f"Modelo tem {len(corrimaos)} IfcRailing. Alturas encontradas nos Psets: {alturas_todas or 'nenhuma'}. "
            f"Nenhum elemento apresentou as duas alturas normativas (0,70m e 0,92m) simultaneamente."
        )

    # ── 5. ESPAÇOS (6.11.1, 7.5) — análise geométrica real ───────────────────
    espacos = []
    n_spaces = contar("IfcSpace")

    if n_spaces > 0:
        try:
            import ifcopenshell.geom
            import numpy as np
            from shapely.geometry import MultiPoint, Point

            geom_settings = ifcopenshell.geom.settings()
            geom_settings.set(geom_settings.USE_WORLD_COORDS, True)

            R_GIRO = 0.75  # raio da área de manobra NBR 9050 item 7.5

            TERMOS_CORREDOR = ["corredor", "circulação", "circulacao", "hall",
                               "acesso", "passagem", "lobby", "foyer", "vestíbulo"]
            TERMOS_SANITARIO = ["banheiro", "sanitário", "sanitario", "wc",
                                "lavabo", "toalete", "vestiário", "vestiario",
                                "banho", "bath", "toilet"]
            TERMOS_ACESSIVEL_SPACE = ["pne", "pcd"]

            for el in ifc.by_type("IfcSpace"):
                d = info_basica(el, "IfcSpace")
                d["LongName"] = getattr(el, "LongName", None)
                ps = todos_psets(el)
                d["Psets"] = ps
                d["Area_m2"] = buscar_prop(ps, "area", "grossarea", "netarea")

                nome_completo = ((d.get("Name") or "") + " " + (d.get("LongName") or "")).lower()
                eh_sanitario_generico = any(t in nome_completo for t in TERMOS_SANITARIO)
                eh_acessivel = any(t in nome_completo for t in TERMOS_ACESSIVEL_SPACE)
                d["tipo_ambiente"] = (
                    "corredor" if any(t in nome_completo for t in TERMOS_CORREDOR)
                    # Só marca como "sanitario" (analisado no item 7.5 — giro 1,50m) se tiver
                    # PNE/PCD no nome. Banheiro comum sem essa tag vira "sanitario_nao_pne"
                    # e fica de fora da verificação de giro/transferência lateral.
                    else "sanitario" if (eh_sanitario_generico and eh_acessivel)
                    else "sanitario_nao_pne" if eh_sanitario_generico
                    else "outro"
                )

                # ── Análise geométrica com ifcopenshell.geom + Shapely ──
                try:
                    shape = ifcopenshell.geom.create_shape(geom_settings, el)
                    verts = np.array(shape.geometry.verts).reshape(-1, 3)
                    z_min = verts[:, 2].min()

                    # Pontos do piso (tolerância 2cm)
                    floor_pts = verts[np.abs(verts[:, 2] - z_min) < 0.02]

                    if len(floor_pts) >= 3:
                        hull = MultiPoint(floor_pts[:, :2]).convex_hull
                        area_geom = round(hull.area, 3)
                        d["area_geometrica_m2"] = area_geom

                        # Bounding box para estimativa de largura (corredores)
                        minx, miny, maxx, maxy = hull.bounds
                        largura_bb  = round(min(maxx - minx, maxy - miny), 3)
                        comprimento_bb = round(max(maxx - minx, maxy - miny), 3)
                        d["largura_estimada_m"]    = largura_bb
                        d["comprimento_estimado_m"] = comprimento_bb

                        # Teste de giro ⌀ 1,50m (NBR 9050 item 7.5)
                        cx, cy = hull.centroid.x, hull.centroid.y
                        circle_centro = Point(cx, cy).buffer(R_GIRO, resolution=64)

                        if hull.contains(circle_centro):
                            d["giro_150_conforme"] = True
                            d["giro_150_status"]   = "Conforme"
                            d["giro_150_nota"]     = f"Círculo ⌀1,50m contido no polígono do ambiente (centróide)"
                        else:
                            # Tenta outras posições (canto, deslocado)
                            encontrou = False
                            for dx, dy in [(0.3, 0), (-0.3, 0), (0, 0.3), (0, -0.3),
                                           (0.5, 0.5), (-0.5, 0.5), (0.5, -0.5), (-0.5, -0.5)]:
                                circle_alt = Point(cx + dx, cy + dy).buffer(R_GIRO, resolution=64)
                                if hull.contains(circle_alt):
                                    encontrou = True
                                    d["giro_150_conforme"] = True
                                    d["giro_150_status"]   = "Conforme"
                                    d["giro_150_nota"]     = f"Círculo ⌀1,50m contido (posição deslocada {dx},{dy}m do centróide)"
                                    break
                            if not encontrou:
                                d["giro_150_conforme"] = False
                                d["giro_150_status"]   = "Não Conforme"
                                d["giro_150_nota"]     = (
                                    f"Círculo ⌀1,50m NÃO cabe no polígono do ambiente. "
                                    f"Área={area_geom:.2f}m² | Largura≈{largura_bb:.2f}m. "
                                    f"Mín. necessário: ⌀1,50m livre de obstruções."
                                )
                    else:
                        d["giro_150_status"] = "Indeterminado"
                        d["giro_150_nota"]   = "Geometria insuficiente para análise"

                except Exception as e_geom:
                    d["giro_150_status"] = "Indeterminado"
                    d["giro_150_nota"]   = f"Erro na extração geométrica: {str(e_geom)[:80]}"

                espacos.append(d)

        except ImportError:
            # Shapely ou ifcopenshell.geom não disponível — fallback básico
            for el in ifc.by_type("IfcSpace"):
                d = info_basica(el, "IfcSpace")
                d["LongName"] = getattr(el, "LongName", None)
                ps = todos_psets(el)
                d["Psets"] = ps
                d["Area_m2"] = buscar_prop(ps, "area", "grossarea", "netarea")
                d["giro_150_status"] = "Indeterminado"
                d["giro_150_nota"]   = "Shapely não disponível — instale: pip install shapely"
                espacos.append(d)

    resultado["elementos"]["IfcSpace"] = espacos[:40]
    resultado["nota_espacos"] = (
        f"Modelo tem {n_spaces} IfcSpace. "
        + (f"Analisados geometricamente: {len(espacos)} espaços com teste de giro ⌀1,50m via Shapely."
           if n_spaces > 0
           else "AUSENTE: não é possível verificar largura de corredores (6.11.1) nem giro de cadeira (7.5) sem IfcSpace. "
                "Recomenda-se exportar Rooms do Revit como IfcSpace com opção 'Export Rooms as IfcSpace'.")
    )

    # ── 6. SANITÁRIOS — classificados por tipo (7.7.2.1, 7.7.1, 7.8, 7.6-7.8) ─
    # IfcFlowTerminal no Revit IFC2X3 contém TUDO: bacias, lavatórios, barras, torneiras
    # Classificação pelo Name (em português, com marca Deca/Celite/Bobrick)

    def get_z_placement(el):
        """
        Extrai coordenada Z RELATIVA ao pavimento — proxy da altura de instalação.
        
        Problema: projetos em coordenadas compartilhadas têm Z global ~700m+.
        Solução: pega só o Z do nível IMEDIATO (RelativePlacement direto),
        ignorando os níveis superiores (pavimento, edifício, terreno).
        Valores plausíveis para equipamentos sanitários: 0.01m a 2.50m.
        """
        try:
            placement = el.ObjectPlacement
            # Pega apenas o placement imediato (relativo ao pavimento)
            if hasattr(placement, "RelativePlacement"):
                rp = placement.RelativePlacement
                if hasattr(rp, "Location") and rp.Location:
                    coords = rp.Location.Coordinates
                    if coords and len(coords) >= 3:
                        z = float(coords[2])
                        # Filtra: só valores plausíveis para altura de equipamento
                        # (entre 1cm e 2,50m — exclui coordenadas globais absurdas)
                        if 0.01 <= z <= 2.50:
                            return round(z, 3)
        except Exception:
            pass
        return None

    TERMOS_BACIA    = ["bacia", "vaso", "toilet", "wc", "p.505", "vogue plus p", "caixa acoplada"]
    TERMOS_LAVAT    = [
        "lavatório", "lavatorio", "lavat",
        "cuba",          # cuba embutir, cuba semiencaixe → sem coluna → Conforme
        "embutir",       # cuba retang. embutir → sem coluna
        "semiencaixe",   # cuba-de-semiencaixe → sem coluna
        "pia", "sink", "basin",
        "l.510", "l.830", "l.733",
    ]
    TERMOS_BARRA    = ["barra apoio", "grab bar", "barra de apoio", "2310.", "2335.", "apoio", "barra "]
    TERMOS_CHUVEIRO = ["chuveiro", "ducha", "shower", "1955", "registro"]

    bacias    = []
    lavatórios = []
    barras    = []
    outros_san = []
    excluidos_sem_pne = []  # fixtures que casaram categoria mas não têm tag PNE/PCD — não entram na auditoria

    for el in ifc.by_type("IfcFlowTerminal"):
        nome  = (getattr(el, "Name", "") or "").lower()
        otype = (getattr(el, "ObjectType", "") or "").lower()
        texto = nome + " " + otype

        d = info_basica(el, "IfcFlowTerminal")
        ps = todos_psets(el)
        d["Psets"] = ps

        # MountingHeight via Psets
        mh = buscar_prop(ps, "mountingheight", "mounting", "instalacao", "installation")
        d["MountingHeight_m"] = mh

        # Z do placement como fallback de altura
        z = get_z_placement(el)
        if z:
            d["Z_placement_m"] = z
            if not mh:
                d["altura_estimada_m"] = z  # proxy para o LLM usar

        pne_ok, pne_fonte = eh_acessivel_pne(el, texto)
        d["pne_pcd_confirmado"] = pne_ok
        if pne_fonte:
            d["pne_pcd_fonte"] = pne_fonte

        if any(t in texto for t in TERMOS_BACIA):
            d["categoria_sanitario"] = "bacia_sanitaria"
            (bacias if pne_ok else excluidos_sem_pne).append(d)
        elif any(t in texto for t in TERMOS_LAVAT):
            d["categoria_sanitario"] = "lavatorio"
            (lavatórios if pne_ok else excluidos_sem_pne).append(d)
        elif any(t in texto for t in TERMOS_BARRA):
            d["categoria_sanitario"] = "barra_apoio"
            (barras if pne_ok else excluidos_sem_pne).append(d)
        elif any(t in texto for t in TERMOS_CHUVEIRO):
            d["categoria_sanitario"] = "chuveiro"
            outros_san.append(d)  # chuveiros/outros não entram no filtro PNE (fora do escopo dos itens 7.x aqui)
        else:
            d["categoria_sanitario"] = "outros"
            outros_san.append(d)

    resultado["elementos"]["Bacias"]    = bacias[:20]
    resultado["elementos"]["Lavatorios"] = lavatórios[:20]
    resultado["elementos"]["BarrasApoio"] = barras[:20]
    resultado["elementos"]["OutrosSanitarios"] = outros_san[:10]
    resultado["nota_sanitarios"] = (
        f"IfcFlowTerminal classificados E marcados como PNE/PCD (analisados nos itens 7.x): "
        f"{len(bacias)} bacias, {len(lavatórios)} lavatórios, {len(barras)} barras de apoio. "
        f"Outros elementos (chuveiros, torneiras, dispensers): {len(outros_san)}. "
        f"Total IfcFlowTerminal no modelo: {contar('IfcFlowTerminal')}. "
        + (
            f"⚠️ {len(excluidos_sem_pne)} bacia(s)/lavatório(s)/barra(s) foram encontrados mas EXCLUÍDOS "
            f"da auditoria por não terem 'PNE' ou 'PCD' no nome (nem no espaço IfcSpace continente) — "
            f"portanto não foram tratados como sanitário acessível. Se isso for inesperado, confira a "
            f"nomenclatura das famílias no Revit (ex: renomear para 'Bacia PNE', 'WC Acessível PCD')."
            if excluidos_sem_pne else
            "Nenhum elemento foi excluído por falta de tag PNE/PCD."
        )
    )

    # ── 7. JANELAS (6.11.3) ──────────────────────────────────────────────────
    janelas = []
    for el in ifc.by_type("IfcWindow"):
        d = info_basica(el, "IfcWindow")
        oh = getattr(el, "OverallHeight", None)
        ow = getattr(el, "OverallWidth", None)
        d["OverallHeight_m"] = round(float(oh), 3) if oh else None
        d["OverallWidth_m"]  = round(float(ow), 3) if ow else None
        ps = todos_psets(el)
        d["Psets"] = ps
        d["SillHeight_m"] = buscar_prop(ps, "sill", "peitoril", "sillheight")
        janelas.append(d)
    resultado["elementos"]["IfcWindow"] = janelas[:40]

    # ── 8. PISOS (6.3.4) ─────────────────────────────────────────────────────
    pisos = []
    for el in ifc.by_type("IfcSlab"):
        nome = (getattr(el, "Name", "") or "").lower()
        otype = (getattr(el, "ObjectType", "") or "").lower()
        if any(t in nome + otype for t in ["rampa", "ramp"]):
            continue
        d = info_basica(el, "IfcSlab")
        ps = todos_psets(el)
        d["Psets"] = ps
        d["Elevation_m"] = buscar_prop(ps, "elevation", "cota", "level")
        pisos.append(d)
    resultado["elementos"]["IfcSlab"] = pisos[:25]

    # ── 9. PAREDES — amostra (fallback para corredores) ───────────────────────
    paredes = []
    for el in list(ifc.by_type("IfcWall"))[:8] + list(ifc.by_type("IfcWallStandardCase"))[:8]:
        d = info_basica(el, el.is_a())
        ps = todos_psets(el)
        d["Width_m"]  = buscar_prop(ps, "width", "thickness", "espessura")
        d["Length_m"] = buscar_prop(ps, "length", "comprimento")
        d["Height_m"] = buscar_prop(ps, "height", "altura")
        paredes.append(d)
    resultado["elementos"]["IfcWall_amostra"] = paredes

    return limpar_nulos(resultado)



# ── Regras da NBR 9050 — agora carregadas de nbr9050_rules.json ─────────────
# ANTES: ~130 linhas de dicionário Python hardcoded (REGRAS_NBR9050).
# DEPOIS: fonte única em JSON, versionável e editável sem tocar neste arquivo.
RULES_PATH = Path(__file__).parent / "nbr9050_rules.json"


@st.cache_data(show_spinner=False)
def carregar_regras(path: Path = RULES_PATH) -> dict:
    """
    Carrega nbr9050_rules.json uma vez por sessão (cache do Streamlit evita
    reler o arquivo do disco a cada interação do usuário).
    """
    if not path.exists():
        st.error(
            f"⚠️ Arquivo de regras não encontrado em `{path}`. "
            "Verifique se nbr9050_rules.json está na raiz do repositório, "
            "junto de nbr9050_app.py."
        )
        return {"itens": []}
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    if "itens" not in payload or not isinstance(payload["itens"], list):
        st.error("⚠️ nbr9050_rules.json malformado: chave 'itens' ausente ou inválida.")
        return {"itens": []}
    return payload


def obter_regras_lista() -> list[dict]:
    """Lista de itens no novo formato (entidades/estrategias aninhados, prompt_llm)."""
    return carregar_regras()["itens"]


def _resumo_elementos(elementos: dict) -> str:
    """
    Gera um resumo estruturado dos elementos extraídos, item por item da NBR.
    Em vez de despejar JSON bruto, apresenta os dados de forma orientada à verificação.
    """
    inv = elementos.get("inventario_modelo", {})
    linhas = []

    # ── Inventário ────────────────────────────────────────────────────────────
    linhas.append("## INVENTÁRIO DO MODELO")
    for entidade, qtd in inv.items():
        linhas.append(f"  {entidade}: {qtd} elementos")

    # ── Notas do modelo ───────────────────────────────────────────────────────
    for k, v in elementos.items():
        if k.startswith("nota_"):
            linhas.append(f"\n⚠️ NOTA — {k.replace('nota_','').upper()}: {v}")

    elems = elementos.get("elementos", {})

    # ── PORTAS ────────────────────────────────────────────────────────────────
    portas = elems.get("IfcDoor", [])
    if portas:
        linhas.append(f"\n## PORTAS (IfcDoor) — amostra de {len(portas)} elementos")
        for p in portas[:15]:
            gid  = p.get("GlobalId","?")
            nome = p.get("Name","?")
            oh   = p.get("OverallHeight_m") or p.get("OverallHeight")
            ow   = p.get("OverallWidth_m")  or p.get("OverallWidth")
            h_str = f"{oh:.3f}m" if oh else "N/D"
            w_str = f"{ow:.3f}m" if ow else "N/D"
            linhas.append(f"  [{gid}] {nome} | Altura={h_str} | Largura={w_str}")
        if len(portas) > 15:
            linhas.append(f"  ... +{len(portas)-15} portas na amostra (ver estatística completa abaixo)")

        est = elementos.get("estatisticas_portas")
        if est and est.get("com_dimensoes"):
            linhas.append(
                f"\n  📊 ESTATÍSTICA SOBRE TODAS AS {est['total']} PORTAS DO MODELO "
                f"(não apenas a amostra acima — use isto para o veredito do item 6.11.2):"
            )
            linhas.append(f"     Menor largura encontrada: {est['largura_min_m']}m [GlobalId: {est['largura_min_globalid']}]")
            linhas.append(f"     Menor altura encontrada: {est['altura_min_m']}m [GlobalId: {est['altura_min_globalid']}]")
            linhas.append(
                f"     Portas com largura < 0,80m: {est['n_nao_conformes_largura']} "
                f"| GlobalIds: {est['globalids_nao_conformes_largura'] or '—'}"
            )
            linhas.append(
                f"     Portas com altura < 2,10m: {est['n_nao_conformes_altura']} "
                f"| GlobalIds: {est['globalids_nao_conformes_altura'] or '—'}"
            )

    # ── ESCADAS ───────────────────────────────────────────────────────────────
    escadas = elems.get("Escadas", [])
    # Filtro robusto — tipo_ifc pode estar em chaves diferentes
    flights = [e for e in escadas if "StairFlight" in str(e.get("tipo_ifc",""))]
    stairs  = [e for e in escadas if e.get("tipo_ifc","") == "IfcStair"]
    linhas.append(f"\n## ESCADAS — {len(stairs)} IfcStair, {len(flights)} IfcStairFlight")
    if flights:
        linhas.append("  ⚠️ IMPORTANTE: RiserHeight_m e TreadLength_m já convertidos de pés→metros (×0.3048)")
        linhas.append("  Use desnivel_m = NumberOfRisers × RiserHeight_m para calcular desnível total")
        for f in flights[:12]:
            gid  = f.get("GlobalId","?")
            nome = (f.get("Name","") or "?")[:50]
            nr   = f.get("NumberOfRisers")
            rh   = f.get("RiserHeight_m")
            tl   = f.get("TreadLength_m")
            dv   = f.get("desnivel_m")
            # Formata com valores reais ou indica ausência
            nr_s = str(int(nr)) if nr else "N/D"
            rh_s = f"{float(rh):.4f}m" if rh else "N/D"
            tl_s = f"{float(tl):.4f}m" if tl else "N/D"
            dv_s = f"{float(dv):.3f}m" if dv else "N/D"
            linhas.append(f"  [{gid}] {nome}")
            linhas.append(f"    NumberOfRisers={nr_s} | RiserHeight_m={rh_s} | TreadLength_m={tl_s} | desnivel_m={dv_s}")
        # Resumo estatístico
        desniveis = [float(f["desnivel_m"]) for f in flights if f.get("desnivel_m")]
        if desniveis:
            linhas.append(f"  RESUMO: desnível mín={min(desniveis):.3f}m | máx={max(desniveis):.3f}m | médio={sum(desniveis)/len(desniveis):.3f}m")
            linhas.append(f"  Todos os {len([d for d in desniveis if d > 0.19])} lances com desnível > 0,19m REQUEREM corrimão.")
    else:
        linhas.append(f"  Nenhum IfcStairFlight com dados — IfcStair no inventário: {inv.get('IfcStair',0)}, IfcStairFlight: {inv.get('IfcStairFlight',0)}")

    # ── RAMPAS ────────────────────────────────────────────────────────────────
    rampas = elems.get("Rampas", [])
    linhas.append(f"\n## RAMPAS — {len(rampas)} elementos")
    if rampas:
        for r in rampas[:5]:
            gid   = r.get("GlobalId","?")
            nome  = r.get("Name","?")
            rise  = r.get("OverallRise_m","N/D")
            run_  = r.get("OverallRun_m","N/D")
            inc   = r.get("inclinacao_pct","?")
            fonte = r.get("fonte_dados_rampa","?")
            linhas.append(f"  [{gid}] {nome} | Rise={rise}m | Run={run_}m | Inclinação={inc}% | Fonte={fonte}")
            if fonte and "ESTIMATIVA" in fonte:
                linhas.append(f"    ⚠️ Rise/Run estimados por bounding box geométrico (Pset não trouxe OverallRise/Run) — conferir manualmente.")
            slope_bruto = r.get("Slope_pset_bruto")
            if slope_bruto is not None:
                linhas.append(f"    Valor bruto de 'Slope' no Pset: {slope_bruto} (unidade não confirmada — pode ser graus; comparar com Inclinação calculada acima antes de usar).")
    else:
        linhas.append("  Nenhuma rampa modelada (IfcRamp/IfcRampFlight ausentes)")

    # ── CORRIMÕES ─────────────────────────────────────────────────────────────
    corrimaos = elems.get("Corrimaos", [])
    linhas.append(f"\n## CORRIMÕES / GUARDA-CORPOS (IfcRailing) — {len(corrimaos)} elementos")
    for c in corrimaos[:8]:
        gid    = c.get("GlobalId","?")
        nome   = c.get("Name","?")
        tipo   = c.get("tipo_elemento","?")
        alturas = c.get("alturas_m", [])
        duplo  = c.get("corrimao_duplo_070_092", False)
        linhas.append(f"  [{gid}] {nome} | Tipo={tipo}")
        linhas.append(f"    Alturas encontradas nos Psets deste elemento: {alturas if alturas else 'N/D'} | Corrimão duplo (0,70m e 0,92m)? {'SIM' if duplo else 'não'}")
    if not corrimaos:
        linhas.append("  Nenhum IfcRailing encontrado")

    # ── SANITÁRIOS: BACIAS ───────────────────────────────────────────────────
    bacias = elems.get("Bacias", [])
    linhas.append(f"\n## BACIAS SANITÁRIAS (IfcFlowTerminal) — {len(bacias)} elementos")
    linhas.append("  VERIFICAR: MountingHeight (Pset) | Z_placement (coordenada Z do modelo) | altura_estimada")
    for b in bacias[:10]:
        gid  = b.get("GlobalId","?")
        nome = (b.get("Name","") or "?")[:55]
        mh   = b.get("MountingHeight_m","—")
        z    = b.get("Z_placement_m","—")
        alt  = b.get("altura_estimada_m","—")
        linhas.append(f"  [{gid}] {nome}")
        linhas.append(f"    MountingHeight={mh} | Z_placement={z}m | altura_estimada={alt}m")
    if bacias:
        linhas.append(f"  NBR 9050 item 7.7.2.1: altura bacia deve ser 0,43m ≤ h ≤ 0,45m (sem assento)")

    # ── SANITÁRIOS: LAVATÓRIOS ───────────────────────────────────────────────
    lavs = elems.get("Lavatorios", [])
    linhas.append(f"\n## LAVATÓRIOS (IfcFlowTerminal) — {len(lavs)} elementos")
    for lv in lavs[:8]:
        gid  = lv.get("GlobalId","?")
        nome = (lv.get("Name","") or "?")[:55]
        mh   = lv.get("MountingHeight_m","—")
        z    = lv.get("Z_placement_m","—")
        linhas.append(f"  [{gid}] {nome} | MountingHeight={mh} | Z={z}m")

    # ── SANITÁRIOS: BARRAS ───────────────────────────────────────────────────
    barras = elems.get("BarrasApoio", [])
    linhas.append(f"\n## BARRAS DE APOIO (IfcFlowTerminal) — {len(barras)} elementos")
    linhas.append("  NBR 9050 item 7.6-7.8: altura instalação ~0,75m | resistência mín 150 kgf")
    for b in barras[:10]:
        gid  = b.get("GlobalId","?")
        nome = (b.get("Name","") or "?")[:55]
        mh   = b.get("MountingHeight_m","—")
        z    = b.get("Z_placement_m","—")
        alt  = b.get("altura_estimada_m","—")
        linhas.append(f"  [{gid}] {nome}")
        linhas.append(f"    MountingHeight={mh} | Z_placement={z}m | altura_estimada={alt}m")

    # ── ESPAÇOS ───────────────────────────────────────────────────────────────
    espacos = elems.get("IfcSpace", [])
    n_total = len(espacos)
    linhas.append(f"\n## ESPAÇOS (IfcSpace) — {n_total} elementos")
    if not espacos:
        linhas.append("  AUSENTE: modelo não exportou IfcSpace")
        linhas.append("  → Corredores (6.11.1) e giro cadeira de rodas (7.5): Indeterminado")
        linhas.append("  → Para habilitar: exportar Rooms do Revit como IfcSpace")
    else:
        # Separa por tipo
        corredores    = [e for e in espacos if e.get("tipo_ambiente") == "corredor"]
        sanitarios_s  = [e for e in espacos if e.get("tipo_ambiente") == "sanitario"]
        sanit_nao_pne = [e for e in espacos if e.get("tipo_ambiente") == "sanitario_nao_pne"]
        outros        = [e for e in espacos if e.get("tipo_ambiente") == "outro"]

        linhas.append(
            f"  Corredores: {len(corredores)} | Sanitários PNE/PCD: {len(sanitarios_s)} | "
            f"Banheiros sem tag PNE/PCD (fora do escopo do item 7.5): {len(sanit_nao_pne)} | Outros: {len(outros)}"
        )

        # Resultados de giro por status — SÓ nos sanitários com tag PNE/PCD (item 7.5 é sobre
        # sanitário acessível, não sobre qualquer banheiro do modelo)
        conformes_giro  = [e for e in sanitarios_s if e.get("giro_150_status") == "Conforme"]
        nconf_giro      = [e for e in sanitarios_s if e.get("giro_150_status") == "Não Conforme"]
        indet_giro      = [e for e in sanitarios_s if e.get("giro_150_status") == "Indeterminado"]

        linhas.append(f"\n  TESTE GIRO ⌀1,50m (NBR 9050 item 7.5) — apenas sanitários PNE/PCD, via Shapely:")
        linhas.append(f"  ✅ Conformes: {len(conformes_giro)} | ❌ Não Conformes: {len(nconf_giro)} | ⚠️ Indeterminados: {len(indet_giro)}")
        if sanit_nao_pne:
            linhas.append(
                f"  ⚠️ {len(sanit_nao_pne)} banheiro(s) encontrados sem 'PNE'/'PCD' no nome — "
                f"EXCLUÍDOS do teste de giro por não serem sanitário acessível designado."
            )

        # Detalha sanitários
        if sanitarios_s:
            linhas.append(f"\n  SANITÁRIOS ACESSÍVEIS:")
            for e in sanitarios_s[:10]:
                gid  = e.get("GlobalId","?")
                nome = (e.get("Name","") or (e.get("LongName","") or "?"))[:40]
                area = e.get("area_geometrica_m2") or e.get("Area_m2","N/D")
                larg = e.get("largura_estimada_m","N/D")
                giro = e.get("giro_150_status","?")
                nota = e.get("giro_150_nota","")
                linhas.append(f"  [{gid}] {nome}")
                linhas.append(f"    Área={area}m² | Largura≈{larg}m | Giro⌀1,50m={giro}")
                if nota:
                    linhas.append(f"    Nota: {nota}")

        # Detalha corredores
        if corredores:
            linhas.append(f"\n  CORREDORES (item 6.11.1 — largura mín por comprimento):")
            for e in corredores[:10]:
                gid  = e.get("GlobalId","?")
                nome = (e.get("Name","") or (e.get("LongName","") or "?"))[:40]
                larg = e.get("largura_estimada_m","N/D")
                comp = e.get("comprimento_estimado_m","N/D")
                # Aplica regra NBR 9050 6.11.1
                if isinstance(larg, (int,float)) and isinstance(comp, (int,float)):
                    if comp <= 4.0:
                        limite = 0.90; regra = "≤4m → mín 0,90m"
                    elif comp <= 10.0:
                        limite = 1.20; regra = "≤10m → mín 1,20m"
                    else:
                        limite = 1.50; regra = ">10m → mín 1,50m"
                    status_corredor = "Conforme" if larg >= limite else "Não Conforme"
                    linhas.append(f"  [{gid}] {nome}")
                    linhas.append(f"    Largura={larg}m | Comp={comp}m | Regra: {regra} | Status: {status_corredor}")
                else:
                    linhas.append(f"  [{gid}] {nome} | Largura={larg}m | Comp={comp}m")

    # ── JANELAS ───────────────────────────────────────────────────────────────
    janelas = elems.get("IfcWindow", [])
    linhas.append(f"\n## JANELAS (IfcWindow) — {len(janelas)} elementos")
    for j in janelas[:5]:
        gid = j.get("GlobalId","?")
        nome = j.get("Name","?")
        sh = j.get("SillHeight_m","N/D")
        h  = j.get("OverallHeight_m","N/D")
        w  = j.get("OverallWidth_m","N/D")
        linhas.append(f"  [{gid}] {nome} | SillHeight={sh} | H={h}m | W={w}m")

    # ── PAREDES (fallback corredores) ─────────────────────────────────────────
    paredes = elems.get("IfcWall_amostra", [])
    if paredes:
        linhas.append(f"\n## PAREDES — amostra ({len(paredes)} de {inv.get('IfcWall',0)+inv.get('IfcWallStandardCase',0)} total)")
        linhas.append("  Use para estimar largura de corredores se IfcSpace ausente")
        for p in paredes[:5]:
            gid = p.get("GlobalId","?")
            nome = p.get("Name","?")
            w = p.get("Width_m","N/D")
            l = p.get("Length_m","N/D")
            linhas.append(f"  [{gid}] {nome} | Esp={w}m | Comp={l}m")

    return "\n".join(linhas)


def build_audit_prompt(elementos: dict, modelo_nome: str) -> str:
    """
    Constrói prompt de auditoria orientado a dados estruturados.
    Em vez de JSON bruto, envia resumo legível por item.
    Regras sempre vêm de nbr9050_rules.json (fonte única) — sem opção de planilha.
    """
    schema = elementos.get("schema", "IFC2X3")
    regras_uso = obter_regras_lista()

    # Instruções por item
    instrucoes = ""
    for r in regras_uso:
        status_possiveis = " | ".join(r.get("status_validacao_possiveis", ["Conforme","Não Conforme","Indeterminado","N/A"]))
        confianca_txt = "SIM — classificação qualitativa/heurística, marque requer_confirmacao_humana=true" if r.get("requer_nivel_confianca") else "não — dado geométrico/objetivo, requer_confirmacao_humana=false"
        instrucoes += f"""
### Item {r['item_nbr']} — {r['subcategoria']}
Verificação: {r['item_verificavel']}
Status possíveis para ESTE item: {status_possiveis}
Requer confirmação humana? {confianca_txt}
Entidade primária: {r['entidades']['primaria']} | Fallback: {r['entidades']['fallback']}
Estratégia primária: {r['estrategias']['primaria']}
Estratégia fallback: {r['estrategias']['fallback']}
Instrução: {r['prompt_llm']}
"""

    # Resumo estruturado dos dados (substitui JSON bruto)
    resumo_dados = _resumo_elementos(elementos)

    n = len(regras_uso)
    return f"""Você é um auditor especialista em acessibilidade arquitetônica — ABNT NBR 9050:2020.
Modelo: "{modelo_nome}" | Schema IFC: {schema}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
REGRAS OBRIGATÓRIAS DE AUDITORIA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. NUNCA retorne "Indeterminado" porque "não encontrou IfcSanitaryTerminal" — 
   este modelo usa IfcFlowTerminal para TODOS os equipamentos sanitários.
   As seções BACIAS, LAVATÓRIOS e BARRAS DE APOIO já estão classificadas para você.

2. Para ESCADAS: use o campo "desnivel_m" já calculado (NumberOfRisers × RiserHeight_m).
   RiserHeight já está convertido de pés para metros. Desnível > 0,19m → corrimão obrigatório.
   Para verificar o corrimão (item 5.4.3): um único IfcRailing pode conter VÁRIAS alturas
   diferentes (campo "alturas_m" — cada valor vem de um sub-Pset do mesmo elemento, ex:
   corrimão inferior + corrimão superior + guarda-corpo). NÃO conclua "sem corrimão duplo"
   só porque o Name do elemento diz "Guarda-corpo" — confira a lista "alturas_m" e o campo
   "corrimao_duplo_070_092" de cada IfcRailing antes de decidir o status.

3. Para CORREDORES sem IfcSpace: use a amostra de paredes para estimar distâncias,
   ou marque "Indeterminado" com recomendação específica de adicionar IfcSpace.

4. Para itens ausentes no modelo (ex: IfcWindow=0): retorne "N/A" com justificativa
   clara de que o elemento não existe no modelo, não "Indeterminado".

5. SEMPRE inclua o GlobalId do elemento mais representativo em cada resultado.

6. Z_placement_m e altura_estimada_m já estão em coordenadas RELATIVAS ao pavimento
   (valores entre 0,01m e 2,50m). Valores como 722m ou 732m foram FILTRADOS.
   Use esses campos para verificar alturas de bacias e barras de apoio.

7. LAVATÓRIOS: "cuba embutir", "cuba retang. embutir", "cuba-de-semiencaixe" são
   lavatórios SEM coluna → Conforme (item 7.8). "Coluna suspensa" → Não Conforme.

8. Gere EXATAMENTE {n} resultados — um por item listado abaixo.

9. Para PORTAS (item 6.11.2): a lista em "PORTAS (IfcDoor)" é só uma AMOSTRA
   ilustrativa. O veredito de conformidade deve usar a seção "ESTATÍSTICA SOBRE
   TODAS AS N PORTAS DO MODELO" — se n_nao_conformes_largura ou
   n_nao_conformes_altura forem > 0, o status é "Não Conforme" (não "Conforme"),
   mesmo que a amostra pareça toda ok. Cite os GlobalIds não conformes listados.

10. Para o item 4.6.6 (maçaneta tipo alavanca): avalie SOMENTE as portas com
    "pne_pcd_confirmado": true (ver "nota_portas_pne"). Portas sem essa tag NÃO
    entram nessa verificação — não as cite como não conformes nem indeterminadas
    para este item. Se nenhuma porta tiver "pne_pcd_confirmado": true, retorne
    status "N/A" para o item 4.6.6, explicando que nenhuma porta foi associada
    a sanitário/ambiente PNE/PCD. O item 6.11.2 (vão livre) continua valendo
    para todas as portas normalmente, independente dessa tag.

11. Status "Parcial": use SOMENTE nos itens cujo "Status possíveis para ESTE
    item" (ver seção ITENS A VERIFICAR) inclua "Parcial". Regra de limiar:
    conte X elementos conformes de Y elementos avaliados. X=Y → "Conforme".
    X=0 → "Não Conforme". 0<X<Y → "Parcial" (nunca arredonde pra Conforme ou
    Não Conforme). Não invente "Parcial" em itens que avaliam um único
    elemento/espaço — nesses, o status continua binário mesmo com esta opção
    disponível de forma geral no sistema.

12. Campo "requer_confirmacao_humana": true nos itens marcados como tal em
    "Requer confirmação humana?" na seção ITENS A VERIFICAR (tipicamente
    itens Qualitativos/Condicionais, como maçaneta e barras de apoio, cuja
    classificação vem de nome/tipo do elemento, não de medição direta).
    false nos itens objetivos/geométricos (ex: vão livre de porta, giro de
    cadeira de rodas).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ITENS A VERIFICAR
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{instrucoes}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DADOS DO MODELO IFC (use estes dados para verificar cada item acima)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{resumo_dados}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FORMATO DE RESPOSTA — JSON VÁLIDO, SEM MARKDOWN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
⚠️ O exemplo abaixo mostra APENAS A ESTRUTURA esperada. Os valores de
"elemento", "valor_encontrado" e "status" são fictícios — NÃO os copie nem
os use como referência de conteúdo. Preencha cada campo com o que você
efetivamente observou nos DADOS DO MODELO IFC acima, item por item.

{{
  "modelo": "{modelo_nome}",
  "schema_ifc": "{schema}",
  "data_auditoria": "{datetime.now().strftime('%d/%m/%Y')}",
  "resultados": [
    {{
      "item_nbr": "<código do item, ex: 6.11.2>",
      "categoria": "<categoria>",
      "subcategoria": "<subcategoria>",
      "elemento": "<contagem/descrição do(s) elemento(s) verificado(s) — preencher com dado real>",
      "globalid": "<GlobalId real do elemento mais representativo>",
      "tipo_ifc": "<classe IFC, ex: IfcDoor>",
      "valor_encontrado": "<valor medido/observado — se Parcial, escreva 'X de Y elementos conformes'>",
      "valor_exigido": "<critério normativo do item>",
      "status": "<use EXATAMENTE uma destas palavras, sem texto extra: Conforme | Parcial | Não Conforme | Indeterminado | N/A>",
      "requer_confirmacao_humana": "<true se o item tem classificação Qualitativa/Condicional (ver 'requer_nivel_confianca' de cada item), senão false>",
      "recomendacao": "<ação concreta — obrigatório se status for Não Conforme, Parcial ou Indeterminado>"
    }}
  ],
  "resumo": {{
    "total": {n},
    "conformes": 0,
    "nao_conformes": 0,
    "indeterminados": 0,
    "na": 0,
    "percentual_conformidade": "0%"
  }},
  "observacoes_gerais": "Análise geral do modelo..."
}}
"""



def _parse_json_robusto(raw: str) -> dict:
    """Tenta parsear JSON, com fallback para JSON truncado."""
    raw_clean = re.sub(r"```json|```", "", raw).strip()
    # Tentativa 1 — JSON completo
    try:
        return json.loads(raw_clean)
    except json.JSONDecodeError:
        pass
    # Tentativa 2 — extrair bloco { ... } mais externo
    try:
        start = raw_clean.index("{")
        # Fecha o JSON truncado adicionando estrutura mínima
        partial = raw_clean[start:]
        # Conta chaves abertas para tentar fechar
        depth = 0
        last_valid = 0
        for i, c in enumerate(partial):
            if c == "{": depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    last_valid = i + 1
                    break
        if last_valid:
            return json.loads(partial[:last_valid])
    except Exception:
        pass
    # Tentativa 3 — extrair resultados parciais com regex e montar estrutura
    try:
        resultados = []
        pattern = r'\{[^{}]*"item_nbr"[^{}]*\}'
        for m in re.finditer(pattern, raw_clean, re.DOTALL):
            try:
                resultados.append(json.loads(m.group()))
            except Exception:
                pass
        if resultados:
            return {
                "modelo": "Extração parcial — resposta truncada",
                "schema_ifc": "—",
                "data_auditoria": datetime.now().strftime("%d/%m/%Y"),
                "resultados": resultados,
                "resumo": {
                    "total": len(resultados),
                    "conformes": sum(1 for r in resultados if "conforme" in r.get("status","").lower() and "não" not in r.get("status","").lower()),
                    "nao_conformes": sum(1 for r in resultados if "não" in r.get("status","").lower() or "nao" in r.get("status","").lower()),
                    "indeterminados": sum(1 for r in resultados if "indet" in r.get("status","").lower()),
                    "na": sum(1 for r in resultados if r.get("status","").lower() == "n/a"),
                    "percentual_conformidade": "parcial"
                },
                "observacoes_gerais": "⚠️ Resposta truncada pelo limite de tokens — resultados parciais exibidos."
            }
    except Exception:
        pass
    raise json.JSONDecodeError("Não foi possível parsear a resposta do modelo.", raw_clean, 0)


def call_anthropic(api_key: str, model: str, prompt: str, temperature: float) -> dict:
    """Call Anthropic API and return parsed JSON result."""
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model=model,
        max_tokens=8192,
        temperature=temperature,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = msg.content[0].text
    return _parse_json_robusto(raw)


def call_gemini(api_key: str, model: str, prompt: str, temperature: float) -> dict:
    """Call Google Gemini API and return parsed JSON result."""
    import google.generativeai as genai
    genai.configure(api_key=api_key)
    m = genai.GenerativeModel(model)
    resp = m.generate_content(
        prompt,
        generation_config={"temperature": temperature, "max_output_tokens": 8192}
    )
    return _parse_json_robusto(resp.text)


def classificar_status(status: str) -> str:
    """
    Classifica a string de status livre devolvida pelo LLM numa das 5 categorias
    canônicas. Tolerante a variações de capitalização/pontuação do modelo
    ("Não conforme", "NÃO CONFORME", "não-conforme" etc. caem todas aqui).

    "Parcial" precisa ser checado ANTES do fallback pra N/A — a palavra não
    contém "não"/"conforme"/"indet", então sem essa checagem explícita todo
    resultado "Parcial" seria contado como "N/A" por engano.

    Fonte única desta regra — usada tanto pelo badge visual (status_badge)
    quanto pelo cálculo do resumo (calcular_resumo), pra garantir que os
    números da tabela e os números do resumo NUNCA divirjam entre si.
    """
    s = (status or "").lower()
    if "parcial" in s:
        return "Parcial"
    if "não" in s or "nao" in s:
        return "Não Conforme"
    if "conforme" in s:
        return "Conforme"
    if "indet" in s:
        return "Indeterminado"
    return "N/A"


def calcular_resumo(resultados: list[dict]) -> dict:
    """
    Calcula o resumo estatístico em PYTHON, de forma determinística —
    em vez de confiar que o LLM soma e divide corretamente.

    ANTES: "percentual_conformidade" vinha inteiro da resposta do modelo
    (o prompt só mostrava um exemplo de formato, "0%", e torcia pra ele
    fazer a conta certa). Reproduzir a mesma auditoria duas vezes podia,
    em teoria, dar dois percentuais diferentes mesmo com os mesmos status.

    DEPOIS: o LLM só precisa classificar cada item (Conforme/Não Conforme/
    Parcial/Indeterminado/N/A) — a contagem e a divisão são sempre feitas
    aqui, logo o mesmo conjunto de status SEMPRE produz o mesmo percentual.

    "Parcial" entra no percentual com peso 0,5 — nem conta como conforme
    (esconderia que parte dos elementos falha), nem como não conforme
    (esconderia que parte já atende). Ex: 4 conformes + 2 parciais + 6
    não conformes, total 12 → (4 + 2×0,5) / 12 = 41,7%.

    Também calcula "percentual_sobre_verificaveis": conformidade excluindo
    itens N/A do denominador. Itens N/A significam "não se aplica a este
    modelo" (ex: item de janela quando o modelo não tem janelas) — incluí-los
    no denominador junto com os itens Indeterminados infla artificialmente
    a sensação de não conformidade. As duas métricas juntas dão um retrato
    mais honesto do que só o percentual bruto.
    """
    total = len(resultados)
    categorias = [classificar_status(r.get("status", "")) for r in resultados]

    conformes      = categorias.count("Conforme")
    parciais       = categorias.count("Parcial")
    nao_conformes  = categorias.count("Não Conforme")
    indeterminados = categorias.count("Indeterminado")
    na             = categorias.count("N/A")

    pontos = conformes + 0.5 * parciais
    verificaveis = total - na
    pct_bruto              = round(pontos / total * 100, 1) if total else 0.0
    pct_sobre_verificaveis = round(pontos / verificaveis * 100, 1) if verificaveis else 0.0

    return {
        "total": total,
        "conformes": conformes,
        "parciais": parciais,
        "nao_conformes": nao_conformes,
        "indeterminados": indeterminados,
        "na": na,
        "percentual_conformidade": f"{pct_bruto}%",
        "percentual_sobre_verificaveis": f"{pct_sobre_verificaveis}%",
    }


def status_badge(status: str) -> str:
    badges = {
        "Conforme":      '<span class="badge badge-conforme">✅ Conforme</span>',
        "Parcial":       '<span class="badge badge-parcial">▲ Parcial</span>',
        "Não Conforme":  '<span class="badge badge-nao">❌ Não Conforme</span>',
        "Indeterminado": '<span class="badge badge-indet">⚠️ Indeterminado</span>',
        "N/A":           '<span class="badge badge-na">— N/A</span>',
    }
    return badges[classificar_status(status)]


AUTORES = "Kevin Dias Quintian &nbsp;·&nbsp; Renata Gomes Rocha &nbsp;·&nbsp; Sergio Rosenboim &nbsp;·&nbsp; Viviane Nishizaki Suzuke &nbsp;·&nbsp; William Felipe dos Santos Moura"
RODAPE_TXT = "Kevin Dias Quintian · Renata Gomes Rocha · Sergio Rosenboim · Viviane Nishizaki Suzuke · William Felipe dos Santos Moura"

def gerar_relatorio_html(resultado: dict, modelo_nome: str) -> str:
    """Generate a self-contained HTML report — Zigurat brand."""
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    resumo = resultado.get("resumo", {})
    itens  = resultado.get("resultados", [])

    # status color map
    def st_color(s):
        sl = s.lower()
        if "parcial" in sl: return "#7c3ac4"
        if "conforme" in sl and "não" not in sl and "nao" not in sl: return "#1ab87a"
        if "não" in sl or "nao" in sl: return "#e03c3c"
        if "indet" in sl: return "#e8920a"
        return "#6b7280"

    rows = ""
    for it in itens:
        gid = it.get("globalid","") or ""
        gid_cell = f'<code style="font-family:\'Courier New\',monospace;font-size:0.73em;background:#eef1f8;border:1px solid #c5cad8;border-radius:4px;padding:2px 7px;color:rgb(28,96,241);letter-spacing:0.02em">{gid}</code>' if gid and gid != "—" else '<span style="color:#aab0be;font-size:0.8em">—</span>'
        st = it.get("status","N/A")
        rec = it.get("recomendacao","") or ""
        confianca_tag = ' <span title="Avaliação qualitativa — recomenda-se confirmação humana" style="cursor:help">🔍</span>' if it.get("requer_confirmacao_humana") else ""
        rows += f"""
        <tr>
          <td><code style="background:#f0f3fb;border-radius:4px;padding:2px 6px;font-size:0.8em;color:rgb(28,96,241)">{it.get('item_nbr','—')}</code>{confianca_tag}</td>
          <td style="color:#3d4252">{it.get('categoria','—')}</td>
          <td style="color:#1a1d26;max-width:200px">{it.get('elemento','—')}</td>
          <td style="color:{st_color(st)};font-weight:700;white-space:nowrap">{st}</td>
          <td style="color:#3d4252;font-family:'Courier New',monospace;font-size:0.8em">{it.get('valor_encontrado','—')}</td>
          <td style="color:#3d4252;font-family:'Courier New',monospace;font-size:0.8em">{it.get('valor_exigido','—')}</td>
          <td style="white-space:nowrap">{gid_cell}</td>
          <td style="color:#6b7280;font-size:0.82em">{it.get('tipo_ifc','—')}</td>
          <td style="color:#b52929;font-size:0.82em">{rec}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Relatório NBR 9050 — {modelo_nome}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: 'Trebuchet MS', Trebuchet, Arial, sans-serif;
    background: #ffffff;
    color: #1a1d26;
    padding: 0;
  }}

  /* ── Header ── */
  .header {{
    background: linear-gradient(120deg, rgb(77,83,99) 0%, rgb(50,56,72) 100%);
    padding: 1.5rem 2.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
  }}
  .header-left h1 {{
    color: rgb(68,205,148);
    font-size: 1.5rem;
    font-weight: 700;
    margin-bottom: 0.2rem;
  }}
  .header-left .sub {{
    color: rgba(255,255,255,0.55);
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.12em;
  }}
  .header-logo {{
    background: #ffffff;
    border-radius: 7px;
    padding: 5px 12px;
    display: inline-flex;
    align-items: center;
    flex-shrink: 0;
  }}
  .header-logo img {{ height: 32px; display: block; }}

  /* ── Body content ── */
  .content {{ padding: 2rem 2.5rem; }}

  /* ── Meta line ── */
  .meta {{
    font-family: 'Courier New', monospace;
    font-size: 0.78rem;
    color: #6b7280;
    margin-bottom: 1.75rem;
    padding-bottom: 1rem;
    border-bottom: 1px solid #e5e7eb;
  }}
  .meta strong {{ color: #1a1d26; }}

  /* ── Section titles ── */
  h2 {{
    font-size: 0.78rem;
    font-weight: 700;
    color: rgb(28,96,241);
    text-transform: uppercase;
    letter-spacing: 0.12em;
    margin: 1.75rem 0 0.85rem 0;
  }}

  /* ── Metric cards ── */
  .resumo {{ display: flex; gap: 0.85rem; flex-wrap: wrap; margin-bottom: 1.5rem; }}
  .card {{
    flex: 1; min-width: 100px;
    background: #f4f6f9;
    border: 1px solid #e5e7eb;
    border-top: 3px solid rgb(68,205,148);
    border-radius: 8px;
    padding: 0.85rem 1rem;
    text-align: center;
  }}
  .card .num {{ font-size: 1.9rem; font-weight: 800; line-height: 1; margin-bottom: 0.2rem; }}
  .card .lbl {{ font-size: 0.6rem; color: #6b7280; text-transform: uppercase; letter-spacing: 0.1em; }}

  /* ── Obs box ── */
  .obs {{
    background: #f4f6f9;
    border-left: 3px solid rgb(68,205,148);
    border-radius: 0 6px 6px 0;
    padding: 1rem 1.25rem;
    font-size: 0.85rem;
    color: #3d4252;
    margin: 0 0 1.5rem 0;
    line-height: 1.65;
  }}

  /* ── GlobalId tip ── */
  .globalid-tip {{
    font-size: 0.75rem;
    color: #6b7280;
    margin-bottom: 0.75rem;
    padding: 0.5rem 0.85rem;
    background: #eef1f8;
    border-radius: 6px;
    border-left: 3px solid rgb(28,96,241);
  }}
  .globalid-tip strong {{ color: rgb(28,96,241); }}

  /* ── Table ── */
  table {{ width: 100%; border-collapse: collapse; font-size: 0.8rem; }}
  thead {{ position: sticky; top: 0; }}
  th {{
    background: rgb(77,83,99);
    color: #ffffff;
    font-size: 0.65rem;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    padding: 10px 12px;
    text-align: left;
    border-bottom: 2px solid rgb(28,96,241);
    white-space: nowrap;
  }}
  th.col-gid {{ color: rgb(68,205,148); }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #f0f1f4; vertical-align: top; }}
  tr:nth-child(even) td {{ background: #fafbfc; }}
  tr:hover td {{ background: rgba(68,205,148,0.06); }}

  /* ── Footer ── */
  .footer {{
    background: rgb(77,83,99);
    padding: 1rem 2.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-top: 2.5rem;
  }}
  .footer-left {{
    font-size: 0.7rem;
    color: rgba(255,255,255,0.5);
    line-height: 1.6;
  }}
  .footer-left strong {{ color: rgb(68,205,148); display: block; margin-bottom: 0.2rem; font-size: 0.72rem; }}
  .footer-left .autores {{ color: rgba(255,255,255,0.7); }}
  .footer-logo {{
    background: #ffffff;
    border-radius: 7px;
    padding: 4px 10px;
    display: inline-flex;
    align-items: center;
  }}
  .footer-logo img {{ height: 20px; display: block; }}
  .footer-right {{
    font-family: 'Courier New', monospace;
    font-size: 0.65rem;
    color: rgba(255,255,255,0.35);
    text-align: right;
  }}
</style>
</head>
<body>

<!-- ── Header ── -->
<div class="header">
  <div class="header-left">
    <h1>&#9855; Relatório de Verificação de Acessibilidade BIM</h1>
    <div class="sub">Verificação Automatizada de Conformidade &nbsp;·&nbsp; ABNT NBR 9050:2020</div>
  </div>
  <div class="header-logo">
    <img src="https://www.e-zigurat.com/images/logo.svg" alt="Zigurat Institute of Technology" />
  </div>
</div>
</div>

<!-- ── Content ── -->
<div class="content">

  <p class="meta">
    Norma: <strong>ABNT NBR 9050:2020</strong> &nbsp;|&nbsp;
    Modelo: <strong>{modelo_nome}</strong> &nbsp;|&nbsp;
    Schema IFC: <strong>{resultado.get('schema_ifc','—')}</strong> &nbsp;|&nbsp;
    Emitido em: <strong>{now}</strong>
  </p>

  <h2>Resumo Executivo</h2>
  <div class="resumo">
    <div class="card"><div class="num" style="color:#1a1d26">{resumo.get('total',0)}</div><div class="lbl">Total</div></div>
    <div class="card"><div class="num" style="color:#1ab87a">{resumo.get('conformes',0)}</div><div class="lbl">Conformes</div></div>
    <div class="card"><div class="num" style="color:#7c3ac4">{resumo.get('parciais',0)}</div><div class="lbl">Parciais</div></div>
    <div class="card"><div class="num" style="color:#e03c3c">{resumo.get('nao_conformes',0)}</div><div class="lbl">Não Conformes</div></div>
    <div class="card"><div class="num" style="color:#e8920a">{resumo.get('indeterminados',0)}</div><div class="lbl">Indeterminados</div></div>
    <div class="card"><div class="num" style="color:#6b7280">{resumo.get('na',0)}</div><div class="lbl">N/A</div></div>
    <div class="card"><div class="num" style="color:rgb(28,96,241)">{resumo.get('percentual_conformidade','—')}</div><div class="lbl">Conformidade (bruta)</div></div>
    <div class="card"><div class="num" style="color:#0c447c">{resumo.get('percentual_sobre_verificaveis','—')}</div><div class="lbl">Conformidade (s/ N/A)</div></div>
  </div>

  <div class="obs">{resultado.get('observacoes_gerais','—')}</div>

  <h2>Resultados Detalhados por Elemento</h2>
  <div class="globalid-tip">
    💡 A coluna <strong>GlobalId</strong> é o identificador único do elemento no IFC — o "CPF" do elemento.
    Use-o no Revit (<em>Manage → Select by ID</em>), no Navisworks ou no BIMcollab para localizar o elemento diretamente no modelo.
    <br>🔍 ao lado do item = avaliação qualitativa (baseada em nome/tipo, não em medição direta) — recomenda-se confirmação humana.
    <br><span style="color:#7c3ac4;font-weight:700">▲ Parcial</span> = alguns dos elementos avaliados atendem ao critério e outros não (ver "Valor Encontrado" para a proporção).
  </div>

  <table>
    <thead>
      <tr>
        <th>Item NBR</th>
        <th>Categoria</th>
        <th>Elemento</th>
        <th>Status</th>
        <th>Valor Encontrado</th>
        <th>Valor Exigido</th>
        <th class="col-gid">&#128273; GlobalId (IFC/Revit)</th>
        <th>Tipo IFC</th>
        <th>Recomendação</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>

</div><!-- /content -->

<!-- ── Footer ── -->
<div class="footer">
  <div class="footer-left">
    <strong>TFM | Grupo 1</strong>
    <span class="autores">{AUTORES}</span>
    <span style="color:rgba(255,255,255,0.3);font-size:0.65rem;margin-top:0.3rem;display:block">
      Gerado automaticamente por IA — verificação manual complementar necessária para itens qualitativos e indeterminados.
    </span>
  </div>
  <div style="display:flex;flex-direction:column;align-items:flex-end;gap:0.5rem">
    <div class="footer-logo">
      <img src="https://www.e-zigurat.com/images/logo.svg" alt="Zigurat" />
    </div>
    <div class="footer-right">Master IA para AEC &nbsp;·&nbsp; {now}</div>
  </div>
</div>

</body>
</html>"""


def gerar_excel(resultado: dict, modelo_nome: str) -> bytes:
    """Generate XLSX report with openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist NBR 9050"

    # Colors
    H_FILL  = PatternFill("solid", fgColor="0A1628")
    SUB_FILL = PatternFill("solid", fgColor="111827")
    CONF_F  = PatternFill("solid", fgColor="064E3B")
    PARC_F  = PatternFill("solid", fgColor="3B1064")
    NAO_F   = PatternFill("solid", fgColor="4C0519")
    INDET_F = PatternFill("solid", fgColor="4B2B06")
    NA_F    = PatternFill("solid", fgColor="1E293B")

    thin = Side(style="thin", color="1F2D45")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.merge_cells("A1:I1")
    ws["A1"] = f"♿ RELATÓRIO DE ACESSIBILIDADE NBR 9050:2020 — {modelo_nome}"
    ws["A1"].font = Font(bold=True, color="00D4AA", size=12, name="Calibri")
    ws["A1"].fill = H_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Schema IFC: {resultado.get('schema_ifc','—')}"
    ws["A2"].font = Font(color="64748B", size=9, italic=True, name="Calibri")
    ws["A2"].fill = SUB_FILL
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Item NBR", "Categoria", "Elemento", "Status", "Valor Encontrado",
               "Valor Exigido", "GlobalId (Revit/IFC)", "Tipo IFC", "Recomendação"]
    widths   = [12, 16, 32, 16, 20, 20, 32, 18, 45]

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = Font(bold=True, color="94A3B8", size=9, name="Calibri")
        cell.fill = SUB_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[chr(64+i)].width = w
    ws.row_dimensions[3].height = 20

    # Data rows
    status_fills = {
        "conforme": CONF_F, "parcial": PARC_F, "não conforme": NAO_F, "nao conforme": NAO_F,
        "indeterminado": INDET_F, "n/a": NA_F
    }
    status_colors = {
        "conforme": "10B981", "parcial": "C084FC", "não conforme": "F87171", "nao conforme": "F87171",
        "indeterminado": "FCD34D", "n/a": "64748B"
    }

    for r, it in enumerate(resultado.get("resultados", []), start=4):
        st_key = it.get("status","").lower()
        fill = status_fills.get(st_key, NA_F)

        values = [
            it.get("item_nbr",""),
            it.get("categoria",""),
            it.get("elemento",""),
            it.get("status",""),
            it.get("valor_encontrado",""),
            it.get("valor_exigido",""),
            it.get("globalid","—"),
            it.get("tipo_ifc",""),
            it.get("recomendacao",""),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(color="E2E8F0", size=9, name="Calibri")
            if c == 4:
                color = status_colors.get(st_key, "E2E8F0")
                cell.font = Font(color=color, bold=True, size=9, name="Calibri")
            if c == 7:
                cell.font = Font(color="60A5FA", size=9, name="Calibri Mono")
        ws.row_dimensions[r].height = 36

    # Summary sheet
    ws2 = wb.create_sheet("Resumo")
    resumo = resultado.get("resumo", {})
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    ws2.append(["Métrica", "Valor"])
    ws2.append(["Modelo", resultado.get("modelo","—")])
    ws2.append(["Schema IFC", resultado.get("schema_ifc","—")])
    ws2.append(["Data Auditoria", resultado.get("data_auditoria", datetime.now().strftime("%d/%m/%Y"))])
    ws2.append(["Total de Itens", resumo.get("total", 0)])
    ws2.append(["✅ Conformes", resumo.get("conformes", 0)])
    ws2.append(["▲ Parciais", resumo.get("parciais", 0)])
    ws2.append(["❌ Não Conformes", resumo.get("nao_conformes", 0)])
    ws2.append(["⚠️ Indeterminados", resumo.get("indeterminados", 0)])
    ws2.append(["— N/A", resumo.get("na", 0)])
    ws2.append(["% Conformidade", resumo.get("percentual_conformidade","—")])
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════
for k, v in {
    "resultado": None,
    "elementos": None,
    "logs": [],
    "running": False,
    "ifc_nome": "",
}.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="padding:1rem 0 1.5rem 0;border-bottom:2px solid rgb(68,205,148);margin-bottom:1rem">
      <img src="https://www.e-zigurat.com/images/logo.svg"
           style="height:28px;display:block;margin-bottom:0.75rem" alt="Zigurat" />
      <div style="font-family:'Trebuchet MS',Trebuchet,sans-serif;font-size:1rem;font-weight:700;color:rgb(28,96,241)">
        &#9855; NBR 9050 Auditor
      </div>
      <div style="font-family:'Trebuchet MS',Trebuchet,sans-serif;font-size:0.65rem;color:rgb(77,83,99);text-transform:uppercase;letter-spacing:0.1em">
        BIM Accessibility Checker
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("**🔑 Provedor de IA**")
    st.markdown("""
    <div style="font-family:'Trebuchet MS',sans-serif;font-size:0.7rem;font-weight:700;
                color:rgb(77,83,99);text-transform:uppercase;letter-spacing:0.1em;
                margin-bottom:0.5rem">
      ⚙️ Configuração
    </div>
    """, unsafe_allow_html=True)

    st.markdown("**Provedor de IA**")
    provider = st.selectbox("Provedor", ["Anthropic (Claude)", "Google (Gemini)"], label_visibility="collapsed")

    st.markdown("""
    <div style="font-family:'Trebuchet MS',sans-serif;font-size:0.78rem;font-weight:600;
                color:#1a1d26;margin-bottom:2px">
      🔑 Chave API
    </div>
    <div style="font-family:'Trebuchet MS',sans-serif;font-size:0.65rem;color:#6b7280;
                margin-bottom:4px">
      Não armazenada • Apenas nesta sessão
    </div>
    """, unsafe_allow_html=True)
    api_key = st.text_input(
        "Chave API",
        type="password",
        placeholder="sk-ant-..." if "Anthropic" in provider else "AIza...",
        help="Sua chave de API. Não é armazenada nem enviada a terceiros.",
        label_visibility="collapsed"
    )

    st.markdown("**🤖 Modelo LLM**")
    if "Anthropic" in provider:
        model_options = [
            "claude-haiku-4-5",
            "claude-sonnet-4-5",
            "claude-opus-4-5",
        ]
        model_labels = {
            "claude-haiku-4-5":  "Claude Haiku 4.5 (rápido, econômico)",
            "claude-sonnet-4-5": "Claude Sonnet 4.5 (balanceado)",
            "claude-opus-4-5":   "Claude Opus 4.5 (máxima qualidade)",
        }
    else:
        model_options = ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-2.0-flash"]
        model_labels = {
            "gemini-1.5-flash": "Gemini 1.5 Flash (rápido)",
            "gemini-1.5-pro": "Gemini 1.5 Pro (balanceado)",
            "gemini-2.0-flash": "Gemini 2.0 Flash (novo)",
        }

    selected_model = st.selectbox(
        "Modelo",
        model_options,
        format_func=lambda x: model_labels.get(x, x),
        label_visibility="collapsed"
    )

    temperature = 0.0  # Fixo em 0.0 — determinístico para auditoria normativa

    st.markdown("---")
    st.markdown("""
    <div style="font-family:'Trebuchet MS',sans-serif;padding-top:0.25rem;line-height:1.8">
      <div style="color:rgb(68,205,148);font-weight:700;font-size:0.72rem;
                  margin-bottom:0.4rem;letter-spacing:0.05em">TFM | Grupo 1</div>
      <div style="font-size:0.65rem;color:rgb(77,83,99)">
        Kevin Dias Quintian<br>
        Renata Gomes Rocha<br>
        Sergio Rosenboim<br>
        Viviane Nishizaki Suzuke<br>
        William Felipe dos Santos Moura
      </div>
      <div style="margin-top:0.6rem;padding-top:0.5rem;
                  border-top:1px solid var(--border);
                  color:#9ca3af;font-size:0.6rem">
        Master IA para AEC &middot; Zigurat Institute of Technology
      </div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero-block">
  <div class="hero-left">
    <div class="hero-title">&#9855; Auditor de Acessibilidade BIM</div>
    <div class="hero-sub">
      Verificação Automatizada de Conformidade &nbsp;·&nbsp;
      <strong style="color:rgba(255,255,255,0.85)">ABNT NBR 9050:2020</strong>
    </div>
  </div>
  <img src="https://www.e-zigurat.com/images/logo.svg"
       style="height:36px;flex-shrink:0"
       alt="Zigurat Institute of Technology" />
</div>
""", unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_upload, tab_resultado, tab_ajuda = st.tabs(["📁 Arquivos & Execução", "📊 Resultados", "❓ Ajuda"])

# ─────────────────────────────────────────────
with tab_upload:

    # ── Upload card — só o IFC (checklist agora vem sempre de nbr9050_rules.json) ──
    st.markdown("""
    <div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.5rem">
      <div class="section-title" style="margin:0">📐 Modelo BIM</div>
      <span style="background:rgb(28,96,241);color:#fff;font-size:0.6rem;
                   font-weight:700;padding:2px 8px;border-radius:10px;
                   letter-spacing:0.05em">OBRIGATÓRIO</span>
    </div>
    <div style="font-size:0.75rem;color:#6b7280;margin-bottom:0.5rem">
      Arquivo IFC exportado do Revit, ArchiCAD ou Vectorworks.
      Suporta schemas <strong>IFC2X3</strong> e <strong>IFC4</strong>.
    </div>
    """, unsafe_allow_html=True)
    ifc_file = st.file_uploader(
        "Arquivo IFC",
        type=["ifc"],
        help="Formato IFC2X3 ou IFC4. Exportado via Revit, ArchiCAD, Vectorworks etc.",
        label_visibility="collapsed"
    )
    if ifc_file:
        st.markdown(f"""
        <div style="background:rgba(68,205,148,0.08);border:1px solid rgba(68,205,148,0.4);
                    border-radius:6px;padding:0.6rem 0.85rem;margin-top:0.5rem;
                    font-size:0.8rem">
          ✅ <strong>{ifc_file.name}</strong>
          <span style="font-family:'Courier New',monospace;color:#6b7280;font-size:0.72rem;margin-left:8px">
            {ifc_file.size / 1024 / 1024:.1f} MB
          </span>
        </div>""", unsafe_allow_html=True)

    n_regras = len(obter_regras_lista())
    st.markdown(f"""
    <div style="background:#f4f6f9;border:1px solid #e5e7eb;border-radius:6px;
                padding:0.6rem 0.85rem;margin-top:0.75rem;font-size:0.75rem;color:#6b7280">
      📋 Checklist NBR carregado de <code>nbr9050_rules.json</code> — {n_regras} itens verificáveis.
    </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ── Stepper — 3 passos (checklist deixou de ser upload, é automático) ────
    step1 = "done" if api_key else "active"
    step2 = "done" if (api_key and ifc_file) else ("active" if api_key else "pending")
    step3 = "active" if (api_key and ifc_file) else "pending"

    def step_dot(state, n):
        colors = {"done": "rgb(68,205,148)", "active": "rgb(28,96,241)", "pending": "#d1d5de"}
        text_c = {"done": "#fff", "active": "#fff", "pending": "#9ca3af"}
        icon   = {"done": "✓", "active": str(n), "pending": str(n)}
        pulse  = 'animation:pulse 1.5s infinite' if state == "active" else ""
        return f"""<div style="width:28px;height:28px;border-radius:50%;
                    background:{colors[state]};color:{text_c[state]};
                    display:flex;align-items:center;justify-content:center;
                    font-size:0.72rem;font-weight:700;flex-shrink:0;{pulse}">
                    {icon[state]}</div>"""

    def step_label(label, sublabel, state):
        c = "rgb(28,96,241)" if state == "done" else ("#1a1d26" if state == "active" else "#9ca3af")
        return f"""<div>
          <div style="font-size:0.8rem;font-weight:700;color:{c}">{label}</div>
          <div style="font-size:0.65rem;color:#6b7280">{sublabel}</div>
        </div>"""

    arrow = '<div style="color:#d1d5de;font-size:0.9rem;padding:0 4px">→</div>'

    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:6px;padding:1rem 1.25rem;
                background:#f4f6f9;border:1px solid #e5e7eb;border-radius:8px;
                margin-bottom:1.5rem;flex-wrap:wrap;gap:8px">
      <div style="display:flex;align-items:center;gap:8px">
        {step_dot(step1,1)}
        {step_label("API Key","Provedor + chave",step1)}
      </div>
      {arrow}
      <div style="display:flex;align-items:center;gap:8px">
        {step_dot(step2,2)}
        {step_label("Modelo IFC","Arquivo .ifc obrigatório",step2)}
      </div>
      {arrow}
      <div style="display:flex;align-items:center;gap:8px">
        {step_dot(step3,3)}
        {step_label("Executar","Iniciar auditoria",step3)}
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── can_run e mensagens de estado ─────────────────────────────────────
    can_run = bool(ifc_file and api_key)

    # Mensagens de estado inline (sem warn-box solta)
    if not api_key and not ifc_file:
        st.markdown("""
        <div class="info-box">
          Complete os passos <strong>① e ②</strong> na barra lateral e acima para habilitar a auditoria.
        </div>""", unsafe_allow_html=True)
    elif not api_key:
        st.markdown('<div class="warn-box">⚠️ Passo ① — Insira sua chave API na barra lateral.</div>', unsafe_allow_html=True)
    elif not ifc_file:
        st.markdown('<div class="warn-box">⚠️ Passo ② — Carregue um arquivo IFC acima.</div>', unsafe_allow_html=True)

    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        run = st.button("▶ Executar Auditoria", disabled=not can_run, use_container_width=True)

    # ── Execution ──────────────────────────────────────────────────────────────
    if run and can_run:
        st.session_state.logs = []
        st.session_state.resultado = None

        log_box = st.empty()
        step_box = st.empty()
        progress_bar = st.progress(0)

        def log(msg: str):
            ts = datetime.now().strftime("%H:%M:%S")
            st.session_state.logs.append(f"[{ts}] {msg}")
            log_content = "\n".join(st.session_state.logs[-20:])
            log_box.markdown(f'<div class="terminal">{log_content}</div>', unsafe_allow_html=True)

        try:
            # Step 1 — Save IFC
            log("🔄 Salvando arquivo IFC temporariamente...")
            progress_bar.progress(10)
            with tempfile.NamedTemporaryFile(suffix=".ifc", delete=False) as tmp:
                tmp.write(ifc_file.read())
                tmp_path = tmp.name
            st.session_state.ifc_nome = ifc_file.name
            log(f"✅ IFC salvo: {ifc_file.name} ({ifc_file.size/1024/1024:.1f} MB)")

            # Step 2 — Extract IFC elements
            log("🔍 Extraindo elementos do modelo IFC (IfcOpenShell)...")
            progress_bar.progress(25)
            elementos = extract_ifc_elements(tmp_path)
            if "error" in elementos:
                st.error(elementos["error"])
                st.stop()
            resumo_ext = {k: len(v) for k, v in elementos.get("elementos", {}).items()}
            log(f"✅ Elementos extraídos: {resumo_ext}")
            log(f"   Schema IFC detectado: {elementos.get('schema','?')}")
            st.session_state.elementos = elementos
            progress_bar.progress(45)

            # Step 3 — Load rules (sempre do JSON — sem upload de checklist)
            n_regras = len(obter_regras_lista())
            log(f"📋 Usando checklist de nbr9050_rules.json ({n_regras} itens).")
            progress_bar.progress(55)

            # Step 4 — Build prompt
            log("📝 Construindo prompt de auditoria...")
            prompt = build_audit_prompt(elementos, ifc_file.name)
            log(f"   Prompt: ~{len(prompt)//4:,} tokens estimados")
            progress_bar.progress(65)

            # Step 5 — Call LLM
            log(f"🤖 Chamando {selected_model} ({provider})...")
            log("   Aguarde — isso pode levar 30–90 segundos...")
            progress_bar.progress(70)

            if "Anthropic" in provider:
                resultado = call_anthropic(api_key, selected_model, prompt, temperature)
            else:
                resultado = call_gemini(api_key, selected_model, prompt, temperature)

            progress_bar.progress(90)
            log(f"✅ Auditoria concluída!")

            # Sobrescreve campos que o Python já conhece com certeza — não faz
            # sentido confiar que o LLM vai ecoar corretamente algo que já está
            # disponível antes mesmo da chamada (mesmo princípio do resumo abaixo).
            # Foi assim que pegamos o bug: numa rodada real, o LLM devolveu
            # "2024-01-15" no lugar da data certa (10/07/2026), ignorando o
            # exemplo que já estava no prompt.
            resultado["modelo"] = ifc_file.name
            resultado["schema_ifc"] = elementos.get("schema", resultado.get("schema_ifc", "—"))
            resultado["data_auditoria"] = datetime.now().strftime("%d/%m/%Y")

            # Recalcula o resumo em Python — determinístico, não depende do LLM
            # ter feito a soma/divisão certa (ver calcular_resumo() para o porquê).
            resultado["resumo"] = calcular_resumo(resultado.get("resultados", []))
            log("🧮 Resumo e metadados recalculados em Python (não dependem do eco do LLM).")

            resumo = resultado["resumo"]
            log(f"   Total: {resumo.get('total',0)} | ✅ {resumo.get('conformes',0)} | ▲ {resumo.get('parciais',0)} | ❌ {resumo.get('nao_conformes',0)} | ⚠️ {resumo.get('indeterminados',0)}")
            log(f"   Conformidade: {resumo.get('percentual_conformidade')} (bruta) | {resumo.get('percentual_sobre_verificaveis')} (sobre itens verificáveis, exclui N/A)")

            st.session_state.resultado = resultado
            progress_bar.progress(100)
            log("🎉 Relatório pronto! Acesse a aba 'Resultados'.")

            # Cleanup
            os.unlink(tmp_path)

        except json.JSONDecodeError as e:
            log(f"❌ Erro ao parsear resposta JSON do modelo: {e}")
            st.error("O modelo não retornou JSON válido. Tente novamente ou ajuste o modelo/temperatura.")
        except Exception as e:
            log(f"❌ Erro: {e}")
            st.error(f"Erro durante a execução: {e}")


# ─────────────────────────────────────────────
with tab_resultado:
    if st.session_state.resultado is None:
        st.markdown("""
        <div style="text-align:center;padding:4rem 2rem;color:#334155">
          <div style="font-size:3rem;margin-bottom:1rem">📊</div>
          <div style="font-family:'Syne',sans-serif;font-size:1.1rem;color:#475569">
            Nenhuma auditoria executada ainda.
          </div>
          <div style="font-size:0.82rem;color:#334155;margin-top:0.5rem">
            Vá para a aba <strong>Arquivos & Execução</strong> e clique em <strong>Executar Auditoria</strong>.
          </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        resultado = st.session_state.resultado
        resumo = resultado.get("resumo", {})
        itens = resultado.get("resultados", [])

        # Metrics
        total   = resumo.get("total", len(itens))
        conf    = resumo.get("conformes", 0)
        parc    = resumo.get("parciais", 0)
        nconf   = resumo.get("nao_conformes", 0)
        indet   = resumo.get("indeterminados", 0)
        na      = resumo.get("na", 0)
        pct     = resumo.get("percentual_conformidade", "—")
        pct_ver = resumo.get("percentual_sobre_verificaveis", "—")

        st.markdown(f"""
        <div class="metric-row">
          <div class="metric-card"><div class="metric-num c-blue">{total}</div><div class="metric-label">Total</div></div>
          <div class="metric-card"><div class="metric-num c-green">{conf}</div><div class="metric-label">Conformes</div></div>
          <div class="metric-card"><div class="metric-num c-purple">{parc}</div><div class="metric-label">Parciais</div></div>
          <div class="metric-card"><div class="metric-num c-red">{nconf}</div><div class="metric-label">Não Conformes</div></div>
          <div class="metric-card"><div class="metric-num c-amber">{indet}</div><div class="metric-label">Indeterminados</div></div>
          <div class="metric-card"><div class="metric-num c-muted">{na}</div><div class="metric-label">N/A</div></div>
          <div class="metric-card"><div class="metric-num c-green">{pct}</div><div class="metric-label">Conformidade (bruta)</div></div>
          <div class="metric-card"><div class="metric-num c-blue">{pct_ver}</div><div class="metric-label">Conformidade (s/ N/A)</div></div>
        </div>
        """, unsafe_allow_html=True)

        # Observações
        obs = resultado.get("observacoes_gerais", "")
        if obs:
            st.markdown(f'<div class="info-box">💬 <strong>Análise Geral:</strong> {obs}</div>', unsafe_allow_html=True)

        st.markdown("---")

        # GlobalId explanation
        st.markdown("""
        <div class="info-box">
          <strong>🔑 Sobre o GlobalId</strong><br>
          O relatório inclui o <code>GlobalId</code> de cada elemento IFC verificado — é o identificador único do elemento no modelo, como um "CPF" do componente BIM.<br>
          <strong>Como usar no Revit:</strong> aba <em>Manage → Inquiry → IFC GUID</em> para localizar o elemento diretamente.
          No <strong>BIMcollab Zoom</strong>, <strong>Solibri</strong> ou <strong>usBIM viewer</strong> (gratuitos), cole o GlobalId no campo de busca para selecionar o elemento instantaneamente.
        </div>
        """, unsafe_allow_html=True)

        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            filter_status = st.multiselect(
                "Filtrar por Status",
                ["Conforme", "Parcial", "Não Conforme", "Indeterminado", "N/A"],
                default=["Conforme", "Parcial", "Não Conforme", "Indeterminado", "N/A"]
            )
        with col_f2:
            cats = sorted(set(it.get("categoria","") for it in itens if it.get("categoria")))
            filter_cat = st.multiselect("Filtrar por Categoria", cats, default=cats)
        with col_f3:
            search_gid = st.text_input("Buscar GlobalId", placeholder="0yScV...", help="Filtra pelo GlobalId do elemento IFC")

        # Table
        itens_filtrados = [
            it for it in itens
            if it.get("status","") in filter_status
            and it.get("categoria","") in filter_cat
            and (not search_gid or search_gid.lower() in it.get("globalid","").lower())
        ]

        st.markdown(f"""
        <div class="section-title">
          📋 Itens Verificados
          <span style="font-weight:400;color:#475569;font-size:0.75rem">— {len(itens_filtrados)} de {len(itens)} itens</span>
        </div>
        """, unsafe_allow_html=True)

        # Build table HTML
        rows_html = ""
        for it in itens_filtrados:
            gid = it.get("globalid", "—")
            gid_chip = f'<span class="globalid" title="GlobalId para filtro no Revit">{gid}</span>' if gid != "—" else "—"
            rows_html += f"""
            <tr>
              <td style="font-family:var(--mono);color:#94a3b8;white-space:nowrap">{it.get('item_nbr','—')}</td>
              <td style="color:#cbd5e1">{it.get('categoria','—')}</td>
              <td style="color:#e2e8f0">{it.get('elemento','—')}</td>
              <td>{status_badge(it.get('status','N/A'))}</td>
              <td style="font-family:var(--mono);font-size:0.78rem;color:#94a3b8">{it.get('valor_encontrado','—')}</td>
              <td style="font-family:var(--mono);font-size:0.78rem;color:#94a3b8">{it.get('valor_exigido','—')}</td>
              <td>{gid_chip}</td>
              <td style="font-family:var(--mono);font-size:0.72rem;color:#475569">{it.get('tipo_ifc','—')}</td>
              <td style="font-size:0.8rem;color:#f87171">{it.get('recomendacao','—') if classificar_status(it.get('status','')) in ('Não Conforme','Parcial','Indeterminado') else '<span style="color:#64748b">—</span>'}</td>
            </tr>"""

        st.markdown(f"""
        <div style="overflow-x:auto">
        <table class="result-table">
          <thead>
            <tr>
              <th>Item NBR</th><th>Categoria</th><th>Elemento</th><th>Status</th>
              <th>Valor Encontrado</th><th>Valor Exigido</th>
              <th>GlobalId 🔍</th><th>Tipo IFC</th><th>Recomendação</th>
            </tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown('<div class="section-title">⬇️ Exportar Relatório</div>', unsafe_allow_html=True)

        col_dl1, col_dl2, col_dl3 = st.columns(3)
        modelo_nome = st.session_state.ifc_nome or "modelo"
        ts = datetime.now().strftime("%Y%m%d_%H%M")

        with col_dl1:
            html_bytes = gerar_relatorio_html(resultado, modelo_nome).encode("utf-8")
            st.download_button(
                "📄 Baixar Relatório HTML",
                data=html_bytes,
                file_name=f"relatorio_nbr9050_{ts}.html",
                mime="text/html",
                use_container_width=True,
            )

        with col_dl2:
            xlsx_bytes = gerar_excel(resultado, modelo_nome)
            if xlsx_bytes:
                st.download_button(
                    "📊 Baixar Checklist XLSX",
                    data=xlsx_bytes,
                    file_name=f"checklist_nbr9050_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with col_dl3:
            json_bytes = json.dumps(resultado, ensure_ascii=False, indent=2).encode("utf-8")
            st.download_button(
                "🗂 Baixar JSON Completo",
                data=json_bytes,
                file_name=f"auditoria_nbr9050_{ts}.json",
                mime="application/json",
                use_container_width=True,
            )

        # JSON expandable
        with st.expander("🔍 Ver JSON bruto da auditoria"):
            st.json(resultado)


# ─────────────────────────────────────────────
with tab_ajuda:
    st.markdown("""
    <div class="section-title">📖 Como usar o Auditor NBR 9050</div>
    """, unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("""
        **1. Configure a IA (barra lateral)**
        - Escolha o provedor: **Anthropic** (Claude) ou **Google** (Gemini)
        - Cole sua chave API
        - Selecione o modelo desejado
        - Ajuste a temperature (0.0–0.2 recomendado para auditoria)

        **2. Carregue o arquivo**
        - **IFC** (obrigatório): arquivo exportado do Revit, ArchiCAD etc.
          - Formatos suportados: IFC2X3, IFC4
          - O sistema detecta o schema automaticamente
        - O **checklist NBR 9050** já vem embutido em `nbr9050_rules.json`,
          na raiz do repositório — não é preciso enviar planilha nenhuma.

        **3. Execute a auditoria**
        - Clique em **Executar Auditoria**
        - Acompanhe o log em tempo real
        - Aguarde 30–90 segundos (depende do modelo e tamanho do IFC)
        """)

    with col_b:
        st.markdown("""
        **4. Analise os resultados**
        - Filtre por status, categoria ou GlobalId
        - O **GlobalId** identifica cada elemento no IFC

        **5. Como usar o GlobalId no Revit**
        - No Revit: `Manage → Select by ID` → cole o GlobalId
        - No Navisworks: filtro por GUID
        - No Solibri / BIMcollab: filtro por GlobalId no modelo IFC
        - No IfcOpenShell: `ifc.by_guid("0yScV...")`

        **6. Exporte os relatórios**
        - **HTML**: relatório visual completo com todos os dados
        - **XLSX**: checklist com formatação por status (conforme/não conforme)
        - **JSON**: dados brutos para integração com outros sistemas
        """)

    st.markdown("---")
    st.markdown("""
    <div class="section-title">♿ Itens NBR 9050:2020 verificados (padrão)</div>
    """, unsafe_allow_html=True)

    # Lê os mesmos 12 itens que o motor de auditoria usa — nbr9050_rules.json
    # ANTES: lista hardcoded "itens_padrao" duplicando (e podendo divergir de) o JSON.
    # DEPOIS: mesma fonte única (obter_regras_lista()) usada em build_audit_prompt().
    itens_padrao = [
        (r["item_nbr"], r["classificacao"], r["subcategoria"], r["item_verificavel"], r["entidades"]["primaria"])
        for r in obter_regras_lista()
    ]

    cat_colors = {"Geométrica": "rgb(28,96,241)", "Condicional": "#e8920a", "Relacional": "#1ab87a", "Qualitativa": "rgb(77,83,99)"}
    rows_help = ""
    for item_nbr, classificacao, cat, desc, entidade in itens_padrao:
        color = cat_colors.get(classificacao, "#6b7280")
        rows_help += f"""
        <tr>
          <td style="font-family:'Courier New',monospace;color:rgb(28,96,241);font-weight:600">{item_nbr}</td>
          <td><span style="color:{color};font-size:0.78rem;font-weight:600">{classificacao}</span></td>
          <td style="color:#1a1d26">{cat}</td>
          <td style="color:#3d4252">{desc}</td>
          <td style="font-family:'Courier New',monospace;font-size:0.72rem;color:#6b7280">{entidade}</td>
        </tr>"""

    st.markdown(f"""
    <table class="result-table">
      <thead>
        <tr><th>Item NBR</th><th>Tipo</th><th>Categoria</th><th>Verificação</th><th>Entidade IFC</th></tr>
      </thead>
      <tbody>{rows_help}</tbody>
    </table>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warn-box" style="margin-top:1.5rem">
    ⚠️ <strong>Limitações conhecidas:</strong>
    Itens qualitativos (maçaneta alavanca, lavatório suspenso) dependem de atributos textuais raramente preenchidos no IFC.
    Itens como espaço de giro (IfcSpace) ficam Indeterminados se o modelo não exportar espaços.
    Verificação manual complementar é sempre recomendada.
    </div>
    """, unsafe_allow_html=True)
